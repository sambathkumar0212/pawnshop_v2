from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.urls import reverse_lazy
from django.contrib import messages
from django.db.models import Q, Sum, Count
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.contrib.auth.decorators import login_required, permission_required
import csv
import openpyxl
from openpyxl.styles import Font, Alignment
from io import BytesIO

from .models import Item, Category, ItemImage
from .forms import ItemForm, CategoryForm, ItemImageForm


class ItemListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Item
    template_name = 'inventory/item_list.html'
    context_object_name = 'items'
    permission_required = 'inventory.view_item'
    paginate_by = 10
    
    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        
        # First filter by organization
        if user.organization:
            queryset = queryset.filter(branch__organization=user.organization)
        
        # Then filter by user's branch if not a superuser
        if not user.is_superuser and user.branch:
            queryset = queryset.filter(branch=user.branch)
        
        # Filter by search query
        search_query = self.request.GET.get('search', '')
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(item_id__icontains=search_query) |
                Q(description__icontains=search_query) |
                Q(customer__first_name__icontains=search_query) |
                Q(customer__last_name__icontains=search_query)
            )
        
        # Filter by category
        category_id = self.request.GET.get('category', '')
        if category_id and category_id.isdigit():
            queryset = queryset.filter(category_id=category_id)
        
        # Filter by status
        status = self.request.GET.get('status', '')
        if status:
            queryset = queryset.filter(status=status)
        
        # Sorting
        sort_by = self.request.GET.get('sort', '-created_at')  # Default sort by newest first
        valid_sort_fields = {
            'name': 'name',
            '-name': '-name',
            'category': 'category__name',
            '-category': '-category__name',
            'status': 'status',
            '-status': '-status',
            'selling_price': 'selling_price',
            '-selling_price': '-selling_price',
            'created_at': 'created_at',
            '-created_at': '-created_at',
            'branch': 'branch__name',
            '-branch': '-branch__name',
        }
        
        if sort_by in valid_sort_fields:
            queryset = queryset.order_by(valid_sort_fields[sort_by])
        else:
            queryset = queryset.order_by('-created_at')  # Default fallback
            
        return queryset.select_related('category', 'branch')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['categories'] = Category.objects.all()
        context['total_items'] = self.get_queryset().count()
        context['available_items'] = self.get_queryset().filter(status='available').count()
        context['pawned_items'] = self.get_queryset().filter(status='pawned').count()
        context['sold_items'] = self.get_queryset().filter(status='sold').count()
        
        # Add search params for maintaining filters during pagination
        context['search_query'] = self.request.GET.get('search', '')
        context['selected_category'] = self.request.GET.get('category', '')
        context['selected_status'] = self.request.GET.get('status', '')
        context['current_sort'] = self.request.GET.get('sort', '-created_at')
        
        return context

    def get(self, request, *args, **kwargs):
        # Check if this is a download request
        format_type = request.GET.get('download')
        if format_type in ['csv', 'excel']:
            return self.download_data(format_type)
        
        # Otherwise, return the normal response
        return super().get(request, *args, **kwargs)
    
    def download_data(self, format_type):
        # Get the filtered queryset (without pagination)
        items = self.get_queryset()
        
        # Define headers
        headers = [
            'Item ID', 'Name', 'Category', 'Description', 'Status',
            'Purchase Price', 'Selling Price', 'Branch', 'Customer',
            'Created Date', 'Created By'
        ]
        
        if format_type == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="items_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
            
            writer = csv.writer(response)
            writer.writerow(headers)
            
            for item in items:
                writer.writerow([
                    item.item_id,
                    item.name,
                    item.category.name if item.category else '',
                    item.description,
                    item.get_status_display(),
                    item.purchase_price,
                    item.selling_price,
                    item.branch.name if item.branch else '',
                    f"{item.customer.first_name} {item.customer.last_name}" if item.customer else '',
                    item.created_at.strftime('%Y-%m-%d %H:%M:%S') if item.created_at else '',
                    f"{item.created_by.first_name} {item.created_by.last_name}" if item.created_by else ''
                ])
            
            return response
        
        elif format_type == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Items"
            
            # Add headers with styling
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Add data
            for row, item in enumerate(items, 2):
                ws.cell(row=row, column=1, value=item.item_id)
                ws.cell(row=row, column=2, value=item.name)
                ws.cell(row=row, column=3, value=item.category.name if item.category else '')
                ws.cell(row=row, column=4, value=item.description)
                ws.cell(row=row, column=5, value=item.get_status_display())
                ws.cell(row=row, column=6, value=float(item.purchase_price) if item.purchase_price else 0)
                ws.cell(row=row, column=7, value=float(item.selling_price) if item.selling_price else 0)
                ws.cell(row=row, column=8, value=item.branch.name if item.branch else '')
                ws.cell(row=row, column=9, value=f"{item.customer.first_name} {item.customer.last_name}" if item.customer else '')
                ws.cell(row=row, column=10, value=item.created_at)
                ws.cell(row=row, column=11, value=f"{item.created_by.first_name} {item.created_by.last_name}" if item.created_by else '')
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="items_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            
            return response


class ItemDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = Item
    template_name = 'inventory/item_detail.html'
    context_object_name = 'item'
    permission_required = 'inventory.view_item'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        item = self.get_object()
        
        # Add loan information if item is pawned
        if item.status == 'pawned' and hasattr(item, 'loan'):
            context['loan'] = item.loan
            
        # Add related items (same category)
        context['related_items'] = Item.objects.filter(
            category=item.category, 
            status='available'
        ).exclude(pk=item.pk)[:4]
        
        # Add item images
        context['images'] = item.images.all()
        
        return context


class ItemCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Item
    template_name = 'inventory/item_form.html'
    form_class = ItemForm
    permission_required = 'inventory.add_item'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        # Pass the user to filter branches
        kwargs['user'] = self.request.user
        return kwargs
    
    def form_valid(self, form):
        # Set the created_by field
        form.instance.created_by = self.request.user
        
        # If user belongs to a branch and no branch selected, use user's branch
        if not form.instance.branch and self.request.user.branch:
            form.instance.branch = self.request.user.branch
            
        response = super().form_valid(form)
        messages.success(self.request, f'Item "{form.instance.name}" has been created successfully.')
        return response
    
    def get_success_url(self):
        return reverse_lazy('item_detail', kwargs={'pk': self.object.pk})


class ItemUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Item
    template_name = 'inventory/item_form.html'
    form_class = ItemForm
    permission_required = 'inventory.change_item'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def form_valid(self, form):
        # Update the modified_by and modified_at fields
        form.instance.modified_by = self.request.user
        form.instance.modified_at = timezone.now()
        
        response = super().form_valid(form)
        messages.success(self.request, f'Item "{form.instance.name}" has been updated successfully.')
        return response
    
    def get_success_url(self):
        return reverse_lazy('item_detail', kwargs={'pk': self.object.pk})


class ItemDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Item
    template_name = 'inventory/item_confirm_delete.html'
    context_object_name = 'item'
    permission_required = 'inventory.delete_item'
    success_url = reverse_lazy('item_list')
    
    def delete(self, request, *args, **kwargs):
        item = self.get_object()
        messages.success(request, f'Item "{item.name}" has been deleted successfully.')
        return super().delete(request, *args, **kwargs)


class CategoryListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Category
    template_name = 'inventory/category_list.html'
    context_object_name = 'categories'
    permission_required = 'inventory.view_category'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Category.objects.annotate(item_count=Count('items'))
        
        # Search functionality
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(description__icontains=search)
            )
        
        # Sorting
        sort_by = self.request.GET.get('sort', 'name')  # Default sort by name
        valid_sort_fields = {
            'name': 'name',
            '-name': '-name',
            'item_count': 'item_count',
            '-item_count': '-item_count',
            'created_at': 'created_at',
            '-created_at': '-created_at',
        }
        
        if sort_by in valid_sort_fields:
            queryset = queryset.order_by(valid_sort_fields[sort_by])
        else:
            queryset = queryset.order_by('name')  # Default fallback
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        context['current_sort'] = self.request.GET.get('sort', 'name')
        return context

    def get(self, request, *args, **kwargs):
        # Check if this is a download request
        format_type = request.GET.get('download')
        if format_type in ['csv', 'excel']:
            return self.download_data(format_type)
        
        # Otherwise, return the normal response
        return super().get(request, *args, **kwargs)
    
    def download_data(self, format_type):
        # Get the filtered queryset (without pagination)
        categories = self.get_queryset()
        
        # Define headers
        headers = [
            'Name', 'Description', 'Item Count', 'Created Date'
        ]
        
        if format_type == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="categories_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
            
            writer = csv.writer(response)
            writer.writerow(headers)
            
            for category in categories:
                writer.writerow([
                    category.name,
                    category.description,
                    category.item_count,
                    category.created_at.strftime('%Y-%m-%d %H:%M:%S') if category.created_at else ''
                ])
            
            return response
        
        elif format_type == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Categories"
            
            # Add headers with styling
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # Add data
            for row, category in enumerate(categories, 2):
                ws.cell(row=row, column=1, value=category.name)
                ws.cell(row=row, column=2, value=category.description)
                ws.cell(row=row, column=3, value=category.item_count)
                ws.cell(row=row, column=4, value=category.created_at)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="categories_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            
            return response


class CategoryCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Category
    template_name = 'inventory/category_form.html'
    form_class = CategoryForm
    permission_required = 'inventory.add_category'
    success_url = reverse_lazy('category_list')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'Category "{form.instance.name}" has been created successfully.')
        return response


class CategoryUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Category
    template_name = 'inventory/category_form.html'
    form_class = CategoryForm
    permission_required = 'inventory.change_category'
    success_url = reverse_lazy('category_list')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, f'Category "{form.instance.name}" has been updated successfully.')
        return response


class CategoryDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Category
    template_name = 'inventory/category_confirm_delete.html'
    context_object_name = 'category'
    permission_required = 'inventory.delete_category'
    success_url = reverse_lazy('category_list')
    
    def delete(self, request, *args, **kwargs):
        category = self.get_object()
        messages.success(request, f'Category "{category.name}" has been deleted successfully.')
        return super().delete(request, *args, **kwargs)


def add_item_image(request, item_id):
    if not request.user.has_perm('inventory.change_item'):
        messages.error(request, "You don't have permission to add images.")
        return redirect('item_detail', pk=item_id)
    
    item = get_object_or_404(Item, id=item_id)
    
    if request.method == 'POST':
        form = ItemImageForm(request.POST, request.FILES)
        if form.is_valid():
            image = form.save(commit=False)
            image.item = item
            image.uploaded_by = request.user
            image.save()
            messages.success(request, "Image added successfully.")
            return redirect('item_detail', pk=item_id)
    else:
        form = ItemImageForm()
    
    return render(request, 'inventory/add_item_image.html', {'form': form, 'item': item})


def delete_item_image(request, image_id):
    if not request.user.has_perm('inventory.change_item'):
        messages.error(request, "You don't have permission to delete images.")
        return JsonResponse({'success': False, 'error': 'Permission denied'})
    
    image = get_object_or_404(ItemImage, id=image_id)
    item_id = image.item.id
    image.delete()
    
    if request.is_ajax():
        return JsonResponse({'success': True})
        
    messages.success(request, "Image deleted successfully.")
    return redirect('item_detail', pk=item_id)


@login_required
@permission_required('inventory.view_item')
def inventory_search(request):
    """Search inventory items and return results"""
    search_query = request.GET.get('q', '')
    category_id = request.GET.get('category', '')
    status = request.GET.get('status', '')
    
    items = Item.objects.all()
    
    # Filter by organization first
    if request.user.organization:
        items = items.filter(branch__organization=request.user.organization)
    
    # Then filter by user's branch if not a superuser
    if not request.user.is_superuser and request.user.branch:
        items = items.filter(branch=request.user.branch)
    
    # Apply search filters
    if search_query:
        items = items.filter(
            Q(name__icontains=search_query) |
            Q(item_id__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(serial_number__icontains=search_query) |
            Q(brand__icontains=search_query) |
            Q(model__icontains=search_query)
        )
    
    # Filter by category if provided
    if category_id and category_id.isdigit():
        items = items.filter(category_id=category_id)
    
    # Filter by status if provided
    if status:
        items = items.filter(status=status)
    
    # Get all categories for the filter dropdown - filter by organization
    if request.user.organization:
        categories = Category.objects.filter(
            items__branch__organization=request.user.organization
        ).distinct()
    else:
        categories = Category.objects.all()
    
    context = {
        'items': items,
        'categories': categories,
        'search_query': search_query,
        'selected_category': category_id,
        'selected_status': status,
        'total_results': items.count()
    }
    
    return render(request, 'inventory/search_results.html', context)
