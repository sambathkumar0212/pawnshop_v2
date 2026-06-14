from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.views import LoginView, LogoutView, PasswordChangeView, PasswordResetView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin, PermissionRequiredMixin
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView, View
from django.urls import reverse_lazy
from django.contrib import messages
from django.db.models import Count, Sum, Q
from django.utils import timezone
from django import forms
from django.http import JsonResponse, Http404, HttpResponseForbidden
from django.contrib.auth import login
from django.core.files.base import ContentFile
from django.conf import settings
import json
import base64
import os
from django.db import connection
from django.db.models.deletion import ProtectedError
from django.utils import timezone
from datetime import datetime
import os
import glob

from .models import CustomUser, Role, UserActivity, Customer, Organization
from .mixins import RoleBranchAccessMixin
from .forms import (UserFaceCreateForm, UserUpdateForm, OrganizationSignupForm, 
                  OrganizationUpdateForm, OrganizationBranchForm, CustomerForm)
from branches.models import Branch
from inventory.models import Item
from transactions.models import Loan, Sale
from .utils import assign_role_to_user
from utils.download_utils import DownloadMixin
from .audit import log_login, log_logout
from utils.default_photos import get_default_person_photo


class CustomLoginView(LoginView):
    template_name = 'accounts/login.html'
    redirect_authenticated_user = True
    
    def form_valid(self, form):
        response = super().form_valid(form)
        # Log login for admin users
        log_login(self.request.user, self.request)
        return response


class CustomLogoutView(LogoutView):
    """Custom logout view that accepts both GET and POST requests"""
    http_method_names = ['get', 'post']  # Allow both GET and POST methods
    
    def dispatch(self, request, *args, **kwargs):
        # Log logout before user is logged out
        if request.user.is_authenticated:
            log_logout(request.user, request)
        return super().dispatch(request, *args, **kwargs)


class AdminOnlyPasswordChangeView(LoginRequiredMixin, PasswordChangeView):
    """Custom password change view - only admins can change their own password"""
    
    def dispatch(self, request, *args, **kwargs):
        # Staff members cannot change their own password
        if request.user.is_authenticated:
            u = request.user
            if not (getattr(u, 'is_pawnshop_admin', False) or getattr(u, 'is_superuser', False) or getattr(u, 'is_staff', False) or getattr(u, 'is_organization_admin', False)):
                messages.error(request, "Staff members cannot change their own password. Please contact an admin.")
                return HttpResponseForbidden("Staff members cannot change their own password.")
        
        return super().dispatch(request, *args, **kwargs)
    
    def form_valid(self, form):
        response = super().form_valid(form)
        
        # Log password change for admin
        from .audit import log_password_change
        log_password_change(
            user=self.request.user,
            changed_by_admin=self.request.user,
            change_type='user_change',
            ip_address=get_request_info(self.request).get('ip_address'),
            description="Admin changed own password"
        )
        
        return response


class AdminOnlyPasswordResetView(PasswordResetView):
    """
    Custom password reset view - only for staff to request password reset
    BUT only admin can click the reset link and confirm.
    Staff can only request, admin verifies and resets.
    """
    template_name = 'accounts/password_reset_request.html'
    email_template_name = 'accounts/password_reset_email.txt'
    subject_template_name = 'accounts/password_reset_subject.txt'
    success_url = reverse_lazy('password_reset_done')
    
    def form_valid(self, form):
        email = form.cleaned_data['email']
        try:
            user = CustomUser.objects.get(email=email)
            if user.is_pawnshop_admin:
                # Admin can use normal password reset
                return super().form_valid(form)
            else:
                # Staff member is requesting password reset
                messages.info(self.request, "Password reset request has been submitted. An admin will review and reset your password.")
                
                # Create a request record for admin to handle
                from .models import AuditTrail
                AuditTrail.objects.create(
                    admin_user=None,
                    change_type='password_change',
                    model_name='CustomUser',
                    object_id=user.id,
                    object_str=str(user),
                    target_user=user,
                    description=f"Password reset request from {user.get_full_name()} ({user.email})"
                )
                
                return redirect('login')
        except CustomUser.DoesNotExist:
            # Security: don't reveal if email exists or not
            pass
        
        return super().form_valid(form)


def get_request_info(request):
    """Extract IP address and user agent from request"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    
    user_agent = request.META.get('HTTP_USER_AGENT', '')
    
    return {
        'ip_address': ip,
        'user_agent': user_agent
    }


class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'home/home.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        today = timezone.now().date()
        
        # Determine branch filter: prefer user's assigned branch if present
        if getattr(user, 'branch', None):
            # User assigned to a branch: restrict dashboard to that branch only
            branches = Branch.objects.filter(id=user.branch.id)
            branch_filter = Q(branch=user.branch)
        elif getattr(user, 'organization', None):
            # Organization-scoped user (no specific branch): show organization's active branches
            branches = Branch.objects.filter(organization=user.organization, is_active=True)
            branch_filter = Q(branch__organization=user.organization)
        else:
            # Superuser or other global users: show all active branches
            branches = Branch.objects.filter(is_active=True)
            branch_filter = Q()
        
        # Get inventory statistics
        context['total_items'] = Item.objects.filter(branch_filter).count()
        context['available_items'] = Item.objects.filter(branch_filter, status='available').count()
        context['pawned_items'] = Item.objects.filter(branch_filter, status='pawned').count()
        
        # Get loan statistics
        context['active_loans'] = Loan.objects.filter(branch_filter, status='active').count()
        context['overdue_loans'] = Loan.objects.filter(
            branch_filter, status='active', due_date__lt=today
        ).count()
        context['loans_due_today'] = Loan.objects.filter(
            branch_filter, status='active', due_date=today
        ).count()
        
        # Get sales statistics
        context['total_sales'] = Sale.objects.filter(
            branch_filter, status='completed', sale_date=today
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        # Get recent loans with all related data - ensure organization filtering
        context['recent_loans'] = Loan.objects.filter(
            branch_filter
        ).select_related(
            'customer'
        ).prefetch_related(
            'loanitem_set',
            'loanitem_set__item'
        ).order_by('-created_at')[:5]
        
        # Get recent sales
        context['recent_sales'] = Sale.objects.filter(
            branch_filter
        ).select_related(
            'customer', 'item'
        ).order_by('-sale_date', '-created_at')[:5]
        
        # Customer statistics - apply same branch/organization filter
        context['customer_count'] = Customer.objects.filter(branch_filter).count()
        context['new_customers_today'] = Customer.objects.filter(branch_filter, created_at__date=today).count()
        
        # Branch information
        context['branches'] = branches
        context['branch_count'] = branches.count()
        
        return context


class SuperAdminOrganizationListView(LoginRequiredMixin, UserPassesTestMixin, DownloadMixin, ListView):
    """Super admin view to list and manage all organizations"""
    model = Organization
    template_name = 'accounts/superadmin/organization_list.html'
    context_object_name = 'organizations'
    paginate_by = 20
    
    # Download configuration
    download_filename = 'organizations'
    download_fields = ['name', 'organization_type', 'contact_email', 'phone_number', 'address', 'created_at']
    download_headers = ['Name', 'Type', 'Email', 'Phone', 'Address', 'Created At']

    def test_func(self):
        return self.request.user.is_superuser
    
    def get_queryset(self):
        queryset = Organization.objects.annotate(
            user_count=Count('users', distinct=True),
            branch_count=Count('branches', distinct=True),
            customer_count=Count('branches__customers', distinct=True)
        ).order_by('-created_at')
        
        # Add search functionality
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(contact_email__icontains=search) |
                Q(owner__username__icontains=search) |
                Q(owner__email__icontains=search)
            )
        
        # Add status filter
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)
            
        # Add plan filter
        plan = self.request.GET.get('plan')
        if plan:
            queryset = queryset.filter(plan=plan)
            
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status_choices'] = Organization.STATUS_CHOICES
        context['plan_choices'] = Organization.PLAN_CHOICES
        context['current_status'] = self.request.GET.get('status', '')
        context['current_plan'] = self.request.GET.get('plan', '')
        context['search_query'] = self.request.GET.get('search', '')
        
        # Add summary statistics
        context['total_organizations'] = Organization.objects.count()
        context['active_organizations'] = Organization.objects.filter(status='active').count()
        context['suspended_organizations'] = Organization.objects.filter(status='suspended').count()
        context['pending_organizations'] = Organization.objects.filter(status='pending').count()
        
        return context


class SuperAdminOrganizationDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    """Super admin detailed view of an organization"""
    model = Organization
    template_name = 'accounts/superadmin/organization_detail.html'
    context_object_name = 'organization'
    
    def test_func(self):
        return self.request.user.is_superuser
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        org = self.object
        
        # Get organization statistics
        context['users'] = org.users.all()[:10]  # Latest 10 users
        context['branches'] = org.branches.all()
        context['total_customers'] = Customer.objects.filter(branch__organization=org).count()
        context['total_loans'] = Loan.objects.filter(branch__organization=org).count() if hasattr(Loan.objects.first(), 'branch') else 0
        
        # Get recent activity
        context['recent_activity'] = UserActivity.objects.filter(
            user__organization=org
        ).select_related('user').order_by('-timestamp')[:10]
        
        return context


class SuperAdminOrganizationSuspendView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Super admin view to suspend an organization"""
    
    def test_func(self):
        return self.request.user.is_superuser
    
    def post(self, request, pk):
        organization = get_object_or_404(Organization, pk=pk)
        
        if organization.status == 'active':
            organization.status = 'suspended'
            organization.save()
            
            # Deactivate all users in the organization
            organization.users.update(is_active=False)
            
            # Log the action
            UserActivity.objects.create(
                user=request.user,
                activity_type='organization_suspended',
                description=f'Suspended organization: {organization.name}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, f'Organization "{organization.name}" has been suspended successfully.')
        else:
            messages.error(request, 'Organization is not active and cannot be suspended.')
            
        return redirect('superadmin_organization_list')


class SuperAdminOrganizationReactivateView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Super admin view to reactivate an organization"""
    
    def test_func(self):
        return self.request.user.is_superuser
    
    def post(self, request, pk):
        organization = get_object_or_404(Organization, pk=pk)
        
        if organization.status == 'suspended':
            organization.status = 'active'
            organization.save()
            
            # Reactivate all users in the organization
            organization.users.update(is_active=True)
            
            # Log the action
            UserActivity.objects.create(
                user=request.user,
                activity_type='organization_reactivated',
                description=f'Reactivated organization: {organization.name}',
                ip_address=request.META.get('REMOTE_ADDR')
            )
            
            messages.success(request, f'Organization "{organization.name}" has been reactivated successfully.')
        else:
            messages.error(request, 'Organization is not suspended and cannot be reactivated.')
            
        return redirect('superadmin_organization_list')


class SuperAdminOrganizationDeleteView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    """Super admin view to delete an organization"""
    model = Organization
    template_name = 'accounts/superadmin/organization_confirm_delete.html'
    success_url = reverse_lazy('superadmin_organization_list')
    
    def test_func(self):
        return self.request.user.is_superuser
    
    def delete(self, request, *args, **kwargs):
        organization = self.get_object()
        org_name = organization.name
        
        # Log the action before deletion
        UserActivity.objects.create(
            user=request.user,
            activity_type='organization_deleted',
            description=f'Permanently deleted organization: {org_name}',
            ip_address=request.META.get('REMOTE_ADDR')
        )
        
        messages.success(request, f'Organization "{org_name}" and all related data has been permanently deleted.')
        return super().delete(request, *args, **kwargs)


class SuperAdminDashboardView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """Super admin dashboard with system overview"""
    template_name = 'accounts/superadmin/dashboard.html'
    
    def test_func(self):
        return self.request.user.is_superuser
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Organization statistics
        context['total_organizations'] = Organization.objects.count()
        context['active_organizations'] = Organization.objects.filter(status='active').count()
        context['suspended_organizations'] = Organization.objects.filter(status='suspended').count()
        context['pending_organizations'] = Organization.objects.filter(status='pending').count()
        
        # User statistics
        context['total_users'] = CustomUser.objects.count()
        context['active_users'] = CustomUser.objects.filter(is_active=True).count()
        context['organization_admins'] = CustomUser.objects.filter(is_organization_admin=True).count()
        
        # Branch statistics
        context['total_branches'] = Branch.objects.count()
        context['active_branches'] = Branch.objects.filter(is_active=True).count()
        
        # Customer statistics
        context['total_customers'] = Customer.objects.count()
        
        # Recent organizations
        context['recent_organizations'] = Organization.objects.order_by('-created_at')[:5]
        
        # Plan distribution
        plan_stats = Organization.objects.values('plan').annotate(count=Count('id'))
        context['plan_distribution'] = {item['plan']: item['count'] for item in plan_stats}
        
        # Recent activity
        context['recent_activity'] = UserActivity.objects.select_related('user').order_by('-timestamp')[:10]
        
        return context


# Add a simple deployment status check function
def check_deployment_status(request):
    """Simple view to check if the deployment is working correctly"""
    from django.http import JsonResponse
    import django
    
    return JsonResponse({
        'status': 'ok',
        'django_version': django.get_version(),
        'database': connection.vendor,
        'app': 'Pawnshop Management System SaaS'
    })


# Add debugging view for branch issue
class DebugBranchView(LoginRequiredMixin, TemplateView):
    """Debug view to investigate branch display issues"""
    template_name = 'accounts/debug_branch.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Find all branches with Tamil in name
        tamil_branches = Branch.objects.filter(name__icontains='Tamil')
        context['tamil_branches'] = tamil_branches
        
        # Find customers associated with Tamil branches
        customers_with_tamil_branch = Customer.objects.select_related('branch').filter(
            branch__name__icontains='Tamil'
        )
        context['customers_with_tamil_branch'] = customers_with_tamil_branch
        
        # Get the specific branch mentioned
        specific_branch = Branch.objects.filter(name='Tamil_R_Gold_Loans_Sam').first()
        context['specific_branch'] = specific_branch
        
        if specific_branch:
            context['customers_in_specific_branch'] = Customer.objects.select_related('branch').filter(
                branch=specific_branch
            )
        
        # Check for any customers without branches
        customers_without_branch = Customer.objects.filter(branch__isnull=True)
        context['customers_without_branch'] = customers_without_branch
        
        return context

# Add missing views that were referenced in URLs but got truncated

class UserListView(LoginRequiredMixin, PermissionRequiredMixin, DownloadMixin, ListView):
    model = CustomUser  # Changed from User to CustomUser
    template_name = 'accounts/user_list.html'
    context_object_name = 'users'
    permission_required = 'accounts.view_customuser'  # Updated permission
    paginate_by = 20
    
    # Download configuration
    download_filename = 'users'
    download_fields = ['username', 'email', 'first_name', 'last_name', 'role__name', 'branch__name', 'is_active', 'date_joined']
    download_headers = ['Username', 'Email', 'First Name', 'Last Name', 'Role', 'Branch', 'Active', 'Date Joined']
    
    def has_permission(self):
        # Allow organization owners/admins to view users
        if hasattr(self.request.user, 'organization') and self.request.user.organization and self.request.user.organization.owner == self.request.user:
            return True
        return super().has_permission()
    
    def get_queryset(self):
        queryset = super().get_queryset()
        user = self.request.user
        
        # If user is an organization admin/owner, only show users for that organization
        if user.organization:
            queryset = queryset.filter(organization=user.organization)
        
        # Search functionality
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(username__icontains=search) |
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search) |
                Q(phone__icontains=search)
            )
        
        # Status filter
        status = self.request.GET.get('status')
        if status == 'active':
            queryset = queryset.filter(is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active=False)
        
        # Role filter
        role = self.request.GET.get('role')
        if role:
            queryset = queryset.filter(role__name=role)
        
        # Sorting
        sort_by = self.request.GET.get('sort', '-date_joined')  # Default sort by newest first
        valid_sort_fields = {
            'username': 'username',
            '-username': '-username',
            'first_name': 'first_name',
            '-first_name': '-first_name',
            'last_name': 'last_name',
            '-last_name': '-last_name',
            'email': 'email',
            '-email': '-email',
            'branch': 'branch__name',
            '-branch': '-branch__name',
            'role': 'role__name',
            '-role': '-role__name',
            'date_joined': 'date_joined',
            '-date_joined': '-date_joined',
            'last_login': 'last_login',
            '-last_login': '-last_login',
        }
        
        if sort_by in valid_sort_fields:
            queryset = queryset.order_by(valid_sort_fields[sort_by])
        else:
            queryset = queryset.order_by('-date_joined')  # Default fallback
            
        return queryset.select_related('branch', 'role')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        context['selected_status'] = self.request.GET.get('status', '')
        context['selected_role'] = self.request.GET.get('role', '')
        context['current_sort'] = self.request.GET.get('sort', '-date_joined')
        return context


class UserCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = CustomUser
    form_class = UserFaceCreateForm
    template_name = 'accounts/user_face_form.html'
    success_url = reverse_lazy('user_list')
    permission_required = 'accounts.add_customuser'
    
    def form_valid(self, form):
        """Handle user creation with password setup"""
        # Don't save yet, we need to set the password
        user = form.save(commit=False)
        
        # Set the password from the form
        password = form.cleaned_data.get('password')
        if password:
            user.set_password(password)
        
        # Save the user
        user.save()
        
        # Handle face image if provided
        if form.cleaned_data.get('face_image'):
            try:
                # Decode and save face image
                face_data = form.cleaned_data.get('face_image')
                if isinstance(face_data, str) and face_data.startswith('data:image'):
                    # Parse base64 image data
                    header, data = face_data.split(',', 1)
                    import base64
                    image_data = base64.b64decode(data)
                    # TODO: Process face encoding for facial recognition
            except Exception as e:
                messages.warning(self.request, f"Face image upload had issues: {str(e)}")
        
        messages.success(self.request, f"User {user.get_full_name()} created successfully.")
        return redirect(self.success_url)


class UserUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = CustomUser
    form_class = UserUpdateForm
    template_name = 'accounts/user_form.html'
    success_url = reverse_lazy('user_list')
    permission_required = 'accounts.change_customuser'


class UserDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = CustomUser
    template_name = 'accounts/user_detail.html'
    context_object_name = 'user_obj'
    permission_required = 'accounts.view_customuser'


class UserDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = CustomUser
    template_name = 'accounts/user_confirm_delete.html'
    context_object_name = 'user_obj'
    success_url = reverse_lazy('user_list')
    permission_required = 'accounts.delete_customuser'


class RoleListView(LoginRequiredMixin, PermissionRequiredMixin, DownloadMixin, ListView):
    model = Role
    template_name = 'accounts/role_list.html'
    context_object_name = 'roles'
    permission_required = 'accounts.view_role'
    paginate_by = 20
    
    # Download configuration
    download_filename = 'roles'
    download_fields = ['name', 'description', 'permissions__name', 'created_at', 'updated_at']
    download_headers = ['Role Name', 'Description', 'Permissions', 'Created At', 'Updated At']


class RoleCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Role
    template_name = 'accounts/role_form.html'
    fields = ['name', 'description', 'permissions']
    success_url = reverse_lazy('role_list')
    permission_required = 'accounts.add_role'


class RoleUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Role
    template_name = 'accounts/role_form.html'
    fields = ['name', 'description', 'permissions']
    success_url = reverse_lazy('role_list')
    permission_required = 'accounts.change_role'


class RoleDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Role
    template_name = 'accounts/role_confirm_delete.html'
    context_object_name = 'role'
    success_url = reverse_lazy('role_list')
    permission_required = 'accounts.delete_role'


class CustomerListView(LoginRequiredMixin, RoleBranchAccessMixin, DownloadMixin, ListView):
    model = Customer
    template_name = 'accounts/customer_list.html'
    context_object_name = 'customers'
    paginate_by = 20
    
    # Download configuration
    download_filename = 'customers'
    download_fields = ['roll_number', 'first_name', 'last_name', 'phone', 'email', 'address', 'city', 'state', 'zip_code', 'id_type', 'id_number', 'branch__name', 'created_at']
    download_headers = ['Roll Number', 'First Name', 'Last Name', 'Phone', 'Email', 'Address', 'City', 'State', 'ZIP Code', 'ID Type', 'ID Number', 'Branch', 'Created At']

    def get_queryset(self):
        queryset = Customer.objects.select_related('branch', 'branch__organization').prefetch_related('loans')
        user = self.request.user
        
        # Apply branch/region access rules
        queryset = self.filter_queryset_by_branches(queryset, branch_field_name='branch')
        # Additionally ensure organization isolation if applicable
        if user.organization:
            queryset = queryset.filter(branch__organization=user.organization)
        
        # Search functionality
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search) |
                Q(phone__icontains=search) |
                Q(id_number__icontains=search) |
                Q(branch__name__icontains=search)
            )
        
        # Filter functionality
        filter_type = self.request.GET.get('filter')
        if filter_type:
            if filter_type == 'active_loans':
                # Filter customers who have active loans
                queryset = queryset.filter(loans__status='active').distinct()
            elif filter_type == 'recent':
                # Filter customers added in the last 30 days
                from datetime import timedelta
                thirty_days_ago = timezone.now().date() - timedelta(days=30)
                queryset = queryset.filter(created_at__date__gte=thirty_days_ago)
        
        # Sorting
        sort_by = self.request.GET.get('sort', '-created_at')  # Default sort by newest first
        valid_sort_fields = {
            'name': 'first_name',
            '-name': '-first_name',
            'email': 'email',
            '-email': '-email',
            'phone': 'phone',
            '-phone': '-phone',
            'city': 'city',
            '-city': '-city',
            'created_at': 'created_at',
            '-created_at': '-created_at',
            'branch': 'branch__name',
            '-branch': '-branch__name',
        }
        
        if sort_by in valid_sort_fields:
            queryset = queryset.order_by(valid_sort_fields[sort_by])
        else:
            queryset = queryset.order_by('-created_at')  # Default fallback
        
        return queryset

    def get(self, request, *args, **kwargs):
        # Check if download is requested
        download_format = request.GET.get('download')
        if download_format:
            if download_format == 'csv':
                return self.download_csv()
            elif download_format == 'excel':
                return self.download_excel()
            elif download_format == 'pdf':
                return self.download_pdf()
        
        return super().get(request, *args, **kwargs)
    
    def get_download_queryset(self):
        """Get queryset for downloads, ordered by created_at"""
        queryset = self.get_queryset()
        # Override sorting to always order by created_at for downloads
        return queryset.order_by('-created_at')
    
    def get_selected_columns(self):
        """Get selected columns from request"""
        columns_param = self.request.GET.get('columns', '')
        if columns_param:
            return columns_param.split(',')
        return None
    
    def filter_row_data(self, row_data, selected_columns=None):
        """Filter row data based on selected columns"""
        if not selected_columns:
            return row_data
        
        column_keys = [
            'roll_number', 'first_name', 'last_name', 'phone', 'email', 'address', 'city',
            'state', 'zip_code', 'id_type', 'id_number', 'branch', 'created_at'
        ]
        
        filtered_row = []
        for i, col_key in enumerate(column_keys):
            if i < len(row_data):
                # Map field names to their positions
                if self.download_fields[i] == col_key or (col_key == 'branch' and self.download_fields[i] == 'branch__name'):
                    if col_key in selected_columns:
                        filtered_row.append(row_data[i])
        return filtered_row
    
    def get_filtered_headers(self):
        """Get filtered headers based on selected columns"""
        selected_columns = self.get_selected_columns()
        if not selected_columns:
            return self.get_download_headers()
        
        headers = []
        for i, field_name in enumerate(self.download_fields):
            col_key = field_name.split('__')[0] if '__' in field_name else field_name
            if col_key == 'branch':
                col_key = 'branch'
            if col_key in selected_columns and i < len(self.download_headers):
                headers.append(self.download_headers[i])
        return headers
    
    def download_csv(self):
        """Export customers as CSV"""
        from django.http import HttpResponse
        import csv
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{self.get_download_filename("csv")}"'
        
        writer = csv.writer(response)
        writer.writerow(self.get_filtered_headers())
        
        # Get data from download-specific queryset
        selected_columns = self.get_selected_columns()
        for row in self.get_download_data_ordered():
            filtered_row = self.filter_row_data(row, selected_columns)
            writer.writerow(filtered_row)
        
        return response
    
    def get_download_data_ordered(self):
        """Get download data with created_at ordering"""
        queryset = self.get_download_queryset()
        fields = self.download_fields
        
        data = []
        for index, obj in enumerate(queryset, start=1):
            row = []
            for field_name in fields:
                try:
                    if field_name == 'roll_number':
                        row.append(str(index))
                        continue
                    # Handle related fields (e.g., branch__name)
                    if '__' in field_name:
                        parts = field_name.split('__')
                        value = obj
                        for part in parts:
                            value = getattr(value, part, None)
                            if value is None:
                                break
                    else:
                        value = getattr(obj, field_name, None)
                    
                    if value is None:
                        value = ''
                    else:
                        value = str(value)
                    row.append(value)
                except AttributeError:
                    row.append('')
            data.append(row)
        return data
    
    def download_excel(self):
        """Export customers as Excel"""
        from django.http import HttpResponse
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Customers"
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        headers = self.get_filtered_headers()
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data - use ordered data
        selected_columns = self.get_selected_columns()
        for row_idx, row_data in enumerate(self.get_download_data_ordered(), 2):
            filtered_row = self.filter_row_data(row_data, selected_columns)
            for col_idx, value in enumerate(filtered_row, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to response
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.get_download_filename("xlsx")}"'
        
        return response
    
    def download_pdf(self):
        """Export customers as PDF with proper text wrapping"""
        from django.http import HttpResponse
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{self.get_download_filename("pdf")}"'
        
        # Use landscape orientation for better table display
        doc = SimpleDocTemplate(
            response, 
            pagesize=landscape(A4),
            rightMargin=0.3*inch,
            leftMargin=0.3*inch,
            topMargin=0.5*inch,
            bottomMargin=0.5*inch
        )
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=20,
            alignment=1,
        )
        
        # Cell text style for wrapping
        cell_style = ParagraphStyle(
            'CellText',
            parent=styles['Normal'],
            fontSize=7,
            leading=9,
            wordWrap='CJK',
        )
        
        # Title
        title = Paragraph("Customers Report", title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Get headers and data (ordered by created_at)
        headers = self.get_filtered_headers()
        raw_data = self.get_download_data_ordered()
        
        # Filter rows based on selected columns
        selected_columns = self.get_selected_columns()
        filtered_data = []
        for row in raw_data:
            filtered_row = self.filter_row_data(row, selected_columns)
            filtered_data.append(filtered_row)
        
        # Define custom column widths for better distribution
        # Adjust based on typical content: Name, Contact Info need more space
        available_width = landscape(A4)[0] - 0.6*inch
        custom_widths = [
            0.6*inch,   # First Name
            0.6*inch,   # Last Name
            0.8*inch,   # Phone
            1.2*inch,   # Email
            1.0*inch,   # Address
            0.5*inch,   # City
            0.5*inch,   # State
            0.5*inch,   # ZIP
            0.5*inch,   # ID Type
            0.7*inch,   # ID Number
            0.8*inch,   # Branch
            0.8*inch,   # Created At
        ]
        
        # Wrap text in Paragraph objects for proper wrapping
        table_data = []
        
        # Add headers
        header_row = [Paragraph(f'<b>{h}</b>', cell_style) for h in headers]
        table_data.append(header_row)
        
        # Add data rows with wrapped text
        for row in filtered_data:
            wrapped_row = []
            for cell in row:
                # Convert to string and wrap in Paragraph
                cell_text = str(cell) if cell else ''
                # Truncate very long text to prevent extreme wrapping
                if len(cell_text) > 50:
                    cell_text = cell_text[:47] + '...'
                wrapped_row.append(Paragraph(cell_text, cell_style))
            table_data.append(wrapped_row)
        
        # Create table with custom column widths
        table = Table(table_data, colWidths=custom_widths, repeatRows=1)
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            
            # Data rows styling
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('TOPPADDING', (0, 1), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f8f8')]),
        ]))
        
        elements.append(table)
        
        # Add timestamp and record count
        elements.append(Spacer(1, 15))
        from datetime import datetime
        timestamp_style = ParagraphStyle(
            'Timestamp',
            parent=styles['Normal'],
            fontSize=8,
            alignment=1,
        )
        record_count = len(filtered_data)
        info_text = f"Total Records: {record_count} | Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        timestamp = Paragraph(info_text, timestamp_style)
        elements.append(timestamp)
        
        doc.build(elements)
        return response
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        context['filter'] = self.request.GET.get('filter', '')  # Changed from selected_filter to filter
        context['current_sort'] = self.request.GET.get('sort', '-created_at')
        
        # Calculate customer statistics
        user = self.request.user
        base_queryset = Customer.objects.all()
        
        # Apply same organization and branch filtering as in get_queryset
        if user.organization:
            base_queryset = base_queryset.filter(branch__organization=user.organization)
        elif user.branch:
            base_queryset = base_queryset.filter(branch=user.branch)
        
        # Calculate statistics
        context['total_customers'] = base_queryset.count()
        context['customers_with_active_loans'] = base_queryset.filter(loans__status='active').distinct().count()
        context['active_customers'] = context['customers_with_active_loans']  # Added for template compatibility
        context['customers_with_loans'] = context['customers_with_active_loans']  # Added for template compatibility
        
        # Recent customers (last 30 days)
        from datetime import timedelta
        thirty_days_ago = timezone.now().date() - timedelta(days=30)
        context['recent_customers'] = base_queryset.filter(created_at__date__gte=thirty_days_ago).count()
        context['new_customers'] = context['recent_customers']  # Added for template compatibility
        
        return context


class CustomerCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Customer
    form_class = CustomerForm  # Use the CustomerForm instead of fields
    template_name = 'accounts/customer_form.html'
    success_url = reverse_lazy('customer_list')
    permission_required = 'accounts.add_customer'
    
    def form_valid(self, form):
        remove_profile_photo = (self.request.POST.get('remove_profile_photo') == '1')
        remove_id_image = (self.request.POST.get('remove_id_image') == '1')

        # Save profile photo from captured/uploaded base64 if provided.
        profile_photo_data = (self.request.POST.get('profile_photo_data') or '').strip()
        if remove_profile_photo:
            form.instance.profile_photo = ''
        elif profile_photo_data:
            form.instance.profile_photo = profile_photo_data
        elif not form.instance.profile_photo:
            # Auto-save default icon photo when creating customer without photo input.
            form.instance.profile_photo = get_default_person_photo()

        if remove_id_image:
            form.instance.id_image = None

        # Set created_by to current user
        form.instance.created_by = self.request.user
        
        messages.success(self.request, f'Customer {form.instance.full_name} has been created successfully!')
        return super().form_valid(form)
    
    def get_form_kwargs(self):
        """Add user context to form"""
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs


class CustomerDetailView(LoginRequiredMixin, RoleBranchAccessMixin, PermissionRequiredMixin, DetailView):
    model = Customer
    template_name = 'accounts/customer_detail.html'
    context_object_name = 'customer'
    permission_required = 'accounts.view_customer'

    def get_object(self, queryset=None):
        obj = super().get_object(queryset=queryset)
        # enforce branch/region access rules
        self.check_object_branch_access(obj, branch_attr='branch')
        return obj

    def has_permission(self):
        # Allow if user has global permission
        user = self.request.user
        try:
            if user.has_perm('accounts.view_customer'):
                return True
        except Exception:
            pass

        # Allow if the customer belongs to one of the user's allowed branches
        try:
            obj = self.get_object()
            allowed = self.get_allowed_branches(user)
            if allowed is None:
                return True
            if obj.branch and allowed.filter(pk=obj.branch.pk).exists():
                return True
        except Exception:
            pass

        # Otherwise fallback to default permission check
        return super().has_permission()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        customer = self.get_object()
        
        # Add active loans with processed photos
        from transactions.views import process_item_photos_for_display
        active_loans = Loan.objects.filter(
            customer=customer, 
            status='active'
        ).select_related('scheme', 'branch').prefetch_related('loanitem_set__item')
        
        # Process photos for each loan
        for loan in active_loans:
            loan.processed_photos = process_item_photos_for_display(loan.item_photos)

        # If customer profile photo is missing, pick the best available loan photo and set it.
        if not (customer.profile_photo and customer.profile_photo.strip()):
            best_photo = None
            # 1) Prefer customer face capture from any loan.
            for loan in active_loans:
                if getattr(loan, 'customer_face_capture', None):
                    best_photo = loan.customer_face_capture
                    break
            # 2) Fall back to first processed item photo from active loans.
            if not best_photo:
                for loan in active_loans:
                    photos = getattr(loan, 'processed_photos', None) or []
                    if photos:
                        best_photo = photos[0]
                        break
            # 3) If no active loan photo, try any loan for this customer.
            if not best_photo:
                all_loans = Loan.objects.filter(customer=customer).order_by('-created_at')
                for loan in all_loans:
                    if getattr(loan, 'customer_face_capture', None):
                        best_photo = loan.customer_face_capture
                        break
                    photos = process_item_photos_for_display(loan.item_photos)
                    if photos:
                        best_photo = photos[0]
                        break

            if best_photo:
                customer.profile_photo = best_photo
                customer.save(update_fields=['profile_photo'])
                # Refresh object for template use in the same request.
                customer.refresh_from_db(fields=['profile_photo'])
        
        context['active_loans'] = active_loans
        
        # Add all loan history
        context['loans'] = Loan.objects.filter(
            customer=customer
        ).select_related('scheme', 'branch').order_by('-issue_date')
        
        # Add items associated with this customer
        context['items'] = Item.objects.filter(
            customer=customer
        ).select_related('branch')
        # Compute expiry/auction notices: loans that are overdue and past grace period
        from django.utils import timezone
        today = timezone.now().date()
        expiry_notices = []
        # consider both active loans past grace period and loans already marked defaulted
        all_loans = context.get('loans', [])
        for loan in all_loans:
            gp_end = getattr(loan, 'grace_period_end', None)
            if loan.status == 'defaulted' or (loan.status == 'active' and loan.is_overdue and gp_end and gp_end <= today):
                expiry_notices.append(loan)

        context['expiry_notices'] = expiry_notices
        
        return context


class CustomerUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Customer
    form_class = CustomerForm  # Use the CustomerForm instead of fields
    template_name = 'accounts/customer_form.html'
    success_url = reverse_lazy('customer_list')
    permission_required = 'accounts.change_customer'
    
    def form_valid(self, form):
        remove_profile_photo = (self.request.POST.get('remove_profile_photo') == '1')
        remove_id_image = (self.request.POST.get('remove_id_image') == '1')

        profile_photo_data = (self.request.POST.get('profile_photo_data') or '').strip()
        if remove_profile_photo:
            form.instance.profile_photo = ''
        elif profile_photo_data:
            form.instance.profile_photo = profile_photo_data

        if remove_id_image:
            form.instance.id_image = None
        messages.success(self.request, f'Customer {form.instance.full_name} has been updated successfully!')
        return super().form_valid(form)
    
    def get_form_kwargs(self):
        """Add user context to form"""
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs


class CustomerDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Customer
    template_name = 'accounts/customer_confirm_delete.html'
    context_object_name = 'customer'
    success_url = reverse_lazy('customer_list')
    permission_required = 'accounts.delete_customer'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['linked_loans'] = self.object.loans.all().order_by('-created_at')
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        try:
            return super().post(request, *args, **kwargs)
        except ProtectedError:
            loan_numbers = list(self.object.loans.values_list('loan_number', flat=True))
            loan_list = ', '.join(loan_numbers[:5])
            if len(loan_numbers) > 5:
                loan_list += ', ...'
            messages.error(
                request,
                f'Customer cannot be deleted because linked loans exist: {loan_list}'
            )
            return redirect('customer_detail', pk=self.object.pk)


class CustomerJsonView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = Customer
    permission_required = 'accounts.view_customer'
    
    def get(self, request, *args, **kwargs):
        customer = self.get_object()
        data = {
            'id': customer.id,
            'name': customer.full_name,
            'phone': customer.phone,
            'email': customer.email or '',
            'address': customer.address or '',
        }
        return JsonResponse(data)


class ProfileView(LoginRequiredMixin, DetailView):
    model = CustomUser
    template_name = 'accounts/profile.html'
    context_object_name = 'user_obj'
    
    def get_object(self):
        return self.request.user


class ProfileUpdateView(LoginRequiredMixin, UpdateView):
    model = CustomUser
    template_name = 'accounts/profile_edit.html'
    fields = ['first_name', 'last_name', 'email', 'phone']
    success_url = reverse_lazy('profile')
    
    def get_object(self):
        return self.request.user


class OrganizationSignupView(CreateView):
    """View for new organization signup (SaaS model)"""
    model = Organization
    form_class = OrganizationSignupForm
    template_name = 'accounts/organization_signup.html'
    success_url = reverse_lazy('login')
    
    def form_valid(self, form):
        """Create organization and first admin user"""
        from django.db import IntegrityError, transaction
        
        try:
            # Use atomic transaction to ensure data consistency
            with transaction.atomic():
                # Get the form data dict which contains both organization and user
                form_data = form.save(commit=False)
                organization = form_data.get('organization')
                user = form_data.get('user')
                
                if not organization or not user:
                    raise ValueError("Form validation failed: organization or user is None")
                
                # Save user first (required for ForeignKey to work)
                user.save()
                
                # Set organization owner and link user to organization
                organization.owner = user
                user.organization = organization
                
                # Save organization
                organization.save()
                
                # Save user again to ensure organization link is persisted
                user.save()
                
                # Add required permissions and groups to organization owner
                from django.contrib.auth.models import Permission, Group
                from django.contrib.contenttypes.models import ContentType
                from branches.models import Branch
                
                # Get or create Organization Admin group
                org_admin_group, created = Group.objects.get_or_create(name='Organization Admin')
                
                # Get branch permissions
                branch_content_type = ContentType.objects.get_for_model(Branch)
                branch_perms = Permission.objects.filter(
                    content_type=branch_content_type,
                    codename__in=['view_branch', 'add_branch', 'change_branch']
                )
                
                # Add organization owner to Organization Admin group
                user.groups.add(org_admin_group)
                
                # Add permissions to user
                user.user_permissions.add(*branch_perms)
                
                # Mark object for consistency with CreateView pattern
                self.object = organization
                
                # Add success message
                messages.success(
                    self.request, 
                    f'Organization "{organization.name}" created successfully! You can now log in.'
                )
                
                return redirect(self.get_success_url())
                
        except IntegrityError as e:
            # Handle duplicate username or other integrity errors
            error_message = str(e).lower()
            
            if 'username' in error_message:
                form.add_error('username', 'This username is already taken. Please choose a different username.')
            elif 'email' in error_message:
                form.add_error('email', 'This email is already registered. Please use a different email or try logging in.')
            elif 'organization' in error_message or 'slug' in error_message:
                form.add_error('organization_name', 'An organization with this name already exists. Please choose a different name.')
            else:
                # Generic error for other integrity issues
                form.add_error(None, 'An error occurred while creating your account. This information may already be registered. Please try different details.')
            
            messages.error(self.request, 'Unable to complete signup. Please check the errors below and try again.')
            return self.form_invalid(form)
        
        except Exception as e:
            # Catch any other unexpected errors
            messages.error(self.request, f'An unexpected error occurred during signup: {str(e)}')
            return self.form_invalid(form)


class OrganizationDashboardView(LoginRequiredMixin, TemplateView):
    """Organization admin dashboard for SaaS"""
    template_name = 'accounts/organization_dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.request.user
        
        if not user.organization:
            messages.error(self.request, "You are not associated with any organization.")
            return context
            
        organization = user.organization
        context['organization'] = organization
        
        # Get branch statistics
        context['branches'] = Branch.objects.filter(organization=organization)
        context['branch_count'] = context['branches'].count()
        
        # Get user statistics
        context['users'] = CustomUser.objects.filter(organization=organization)
        context['user_count'] = context['users'].count()
        
        # Get customer statistics
        customers = Customer.objects.filter(branch__organization=organization)
        context['customers'] = customers
        context['customer_count'] = context['customers'].count()
        
        # Get subscription details
        context['subscription_active'] = organization.is_subscription_active()
        context['subscription_end'] = organization.subscription_end
        context['plan'] = organization.plan
        
        # Check if limits are being approached
        context['approaching_branch_limit'] = context['branch_count'] >= organization.max_branches * 0.8
        context['approaching_user_limit'] = context['user_count'] >= organization.max_users * 0.8
        context['approaching_customer_limit'] = context['customer_count'] >= organization.max_customers * 0.8
        
        return context
    
    def dispatch(self, request, *args, **kwargs):
        """Ensure user belongs to an organization"""
        if not request.user.organization:
            messages.error(request, "You are not associated with any organization.")
            return redirect('dashboard')
        return super().dispatch(request, *args, **kwargs)


class OrganizationUpdateView(LoginRequiredMixin, UpdateView):
    """View for updating organization details"""
    model = Organization
    form_class = OrganizationUpdateForm
    template_name = 'accounts/organization_form.html'
    success_url = reverse_lazy('organization_dashboard')
    
    def get_object(self):
        return self.request.user.organization


class OrganizationBranchCreateView(LoginRequiredMixin, CreateView):
    """View for creating a new branch within an organization"""
    model = Branch
    form_class = OrganizationBranchForm
    template_name = 'accounts/branch_form.html'
    success_url = reverse_lazy('organization_dashboard')


class OrganizationUserCreateView(LoginRequiredMixin, CreateView):
    """View for creating a new user within an organization"""
    model = CustomUser
    form_class = UserFaceCreateForm
    template_name = 'accounts/user_face_form.html'
    success_url = reverse_lazy('organization_dashboard')
    
    def form_valid(self, form):
        """Handle user creation and link to organization"""
        # Don't save yet, we need to set the password and organization
        user = form.save(commit=False)
        
        # Set the password from the form
        password = form.cleaned_data.get('password')
        if password:
            user.set_password(password)
        else:
            # Generate a temporary password if none provided
            from django.utils.crypto import get_random_string
            temp_password = get_random_string(12)
            user.set_password(temp_password)
        
        # Link user to the current user's organization
        if self.request.user.organization:
            user.organization = self.request.user.organization
        else:
            messages.error(self.request, "You are not associated with any organization.")
            return redirect(self.success_url)
        
        # Save the user
        user.save()
        
        messages.success(self.request, f"User {user.get_full_name()} created and linked to your organization.")
        return redirect(self.success_url)


class FaceEnrollmentView(LoginRequiredMixin, TemplateView):
    """View for Face ID enrollment"""
    template_name = 'accounts/face_enrollment.html'


class FaceLoginView(TemplateView):
    """View for Face ID login"""
    template_name = 'accounts/face_login.html'


class SubscriptionPlansView(LoginRequiredMixin, TemplateView):
    """View subscription plans"""
    template_name = 'accounts/subscription_plans.html'


class SubscriptionUpgradeView(LoginRequiredMixin, View):
    """Handle subscription upgrades"""
    def post(self, request, plan):
        # Handle subscription upgrade logic
        messages.success(request, f'Subscription upgraded to {plan}!')
        return redirect('organization_dashboard')


class ToggleSubscriptionAutoRenewView(LoginRequiredMixin, View):
    """Toggle auto-renewal for subscription"""
    def post(self, request):
        org = request.user.organization
        if org:
            org.auto_renew = not org.auto_renew
            org.save()
            status = "enabled" if org.auto_renew else "disabled"
            messages.success(request, f'Auto-renewal {status}.')
        return redirect('organization_dashboard')


class OrganizationListView(LoginRequiredMixin, PermissionRequiredMixin, DownloadMixin, ListView):
    model = Organization
    template_name = 'accounts/organization_list.html'
    context_object_name = 'organizations'
    permission_required = 'accounts.view_organization'
    paginate_by = 20
    
    # Download configuration
    download_filename = 'organizations'
    download_fields = ['name', 'address', 'phone', 'email', 'is_active', 'created_at']
    download_headers = ['Name', 'Address', 'Phone', 'Email', 'Active', 'Created At']


class BranchListView(LoginRequiredMixin, PermissionRequiredMixin, DownloadMixin, ListView):
    model = Branch
    template_name = 'accounts/branch_list.html'
    context_object_name = 'branches'
    permission_required = 'accounts.view_branch'
    paginate_by = 20
    
    # Download configuration
    download_filename = 'branches'
    download_fields = ['name', 'address', 'phone', 'email', 'manager__username', 'is_active', 'created_at']
    download_headers = ['Name', 'Address', 'Phone', 'Email', 'Manager', 'Active', 'Created Date']
