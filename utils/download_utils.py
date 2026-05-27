import csv
import json
from datetime import datetime
from io import BytesIO
from django.http import HttpResponse
from django.template.loader import get_template
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views.generic import View
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch


class DownloadMixin:
    """
    Mixin to add download functionality to ListView classes
    """
    download_filename = None
    download_fields = None
    download_headers = None
    
    def get_download_filename(self, format_type):
        """Generate filename for downloads"""
        if self.download_filename:
            base_name = self.download_filename
        else:
            base_name = self.model._meta.verbose_name_plural.replace(' ', '_').lower()
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        return f"{base_name}_{timestamp}.{format_type}"
    
    def get_download_fields(self):
        """Get fields to include in download"""
        if self.download_fields:
            return self.download_fields
        
        # Default to all model fields except sensitive ones
        exclude_fields = ['password', 'token', 'secret', 'key']
        fields = []
        for field in self.model._meta.get_fields():
            if hasattr(field, 'name') and not any(ex in field.name.lower() for ex in exclude_fields):
                if not field.many_to_many and not field.one_to_many:
                    fields.append(field.name)
        return fields
    
    def get_download_headers(self):
        """Get headers for download"""
        if self.download_headers:
            return self.download_headers
        
        fields = self.get_download_fields()
        headers = []
        for field_name in fields:
            try:
                field = self.model._meta.get_field(field_name)
                headers.append(field.verbose_name.title())
            except:
                headers.append(field_name.replace('_', ' ').title())
        return headers
    
    def get_download_data(self):
        """Get data for download"""
        queryset = self.get_queryset()
        fields = self.get_download_fields()
        
        data = []
        for obj in queryset:
            row = []
            for field_name in fields:
                try:
                    value = getattr(obj, field_name)
                    if hasattr(value, 'all'):  # Many-to-many field
                        value = ', '.join(str(v) for v in value.all())
                    elif value is None:
                        value = ''
                    else:
                        value = str(value)
                    row.append(value)
                except AttributeError:
                    row.append('')
            data.append(row)
        return data


@method_decorator(login_required, name='dispatch')
class DownloadView(View):
    """
    Generic download view that can export data in CSV, Excel, or PDF format
    """
    
    def get(self, request, *args, **kwargs):
        format_type = request.GET.get('format', 'csv').lower()
        
        # Get the list view class from the referring page
        list_view_class = self.get_list_view_class()
        if not list_view_class:
            return HttpResponse("Invalid download request", status=400)
        
        # Create an instance of the list view
        list_view = list_view_class()
        list_view.request = request
        list_view.args = args
        list_view.kwargs = kwargs
        
        if format_type == 'csv':
            return self.export_csv(list_view)
        elif format_type == 'excel':
            return self.export_excel(list_view)
        elif format_type == 'pdf':
            return self.export_pdf(list_view)
        else:
            return HttpResponse("Unsupported format", status=400)
    
    def get_list_view_class(self):
        """Override this method in subclasses to return the appropriate ListView class"""
        return None
    
    def export_csv(self, list_view):
        """Export data as CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{list_view.get_download_filename("csv")}"'
        
        writer = csv.writer(response)
        writer.writerow(list_view.get_download_headers())
        
        for row in list_view.get_download_data():
            writer.writerow(row)
        
        return response
    
    def export_excel(self, list_view):
        """Export data as Excel"""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = list_view.model._meta.verbose_name_plural.title()
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        headers = list_view.get_download_headers()
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row_idx, row_data in enumerate(list_view.get_download_data(), 2):
            for col_idx, value in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to response
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{list_view.get_download_filename("xlsx")}"'
        
        return response
    
    def export_pdf(self, list_view):
        """Export data as PDF"""
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{list_view.get_download_filename("pdf")}"'
        
        # Create PDF document
        doc = SimpleDocTemplate(response, pagesize=A4)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1,  # Center alignment
        )
        
        # Title
        title = Paragraph(
            f"{list_view.model._meta.verbose_name_plural.title()} Report",
            title_style
        )
        elements.append(title)
        elements.append(Spacer(1, 20))
        
        # Table data
        headers = list_view.get_download_headers()
        data = [headers] + list_view.get_download_data()
        
        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        elements.append(table)
        
        # Add timestamp
        timestamp = Paragraph(
            f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            styles['Normal']
        )
        elements.append(Spacer(1, 30))
        elements.append(timestamp)
        
        doc.build(elements)
        return response


def create_download_buttons_context():
    """Helper function to create download buttons context for templates"""
    return {
        'download_formats': [
            {'format': 'csv', 'label': 'CSV', 'icon': 'fa-file-csv'},
            {'format': 'excel', 'label': 'Excel', 'icon': 'fa-file-excel'},
            {'format': 'pdf', 'label': 'PDF', 'icon': 'fa-file-pdf'},
        ]
    }
