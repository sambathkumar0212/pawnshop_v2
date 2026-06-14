from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, View
from django.urls import reverse_lazy, reverse
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.utils import timezone
from django.http import Http404, HttpResponse, JsonResponse
from django.template.loader import get_template
from io import BytesIO
import csv
from .models import Loan, Payment, LoanExtension, Sale
from accounts.mixins import RoleBranchAccessMixin
from .forms import LoanForm, SaleForm, LoanExtensionForm
from .utils import ManagerPermissionMixin
from django.db.models import Q
from num2words import num2words
from django.core.files.base import ContentFile
from decimal import Decimal
import logging
import base64
import json
import ast
import os
import shutil
import subprocess
import tempfile
import re
from django.conf import settings
from django.db.utils import OperationalError, ProgrammingError
from urllib.parse import urlparse
from utils.download_utils import DownloadMixin
from utils.translation import translate_text


def translate_text_for_pdf(text, target_lang='ta', timeout=5):
    """Translate a short piece of text for PDF output with timeout and fallback."""
    if not text:
        return ''

    if target_lang == 'en':
        return text

    # Skip translation if requests would take too long
    try:
        import requests
        response = requests.get(
            'https://translate.googleapis.com/translate_a/single',
            params={
                'client': 'gtx',
                'sl': 'en',
                'tl': target_lang,
                'dt': 't',
                'q': text,
            },
            headers={
                'User-Agent': 'Mozilla/5.0',
                'Accept': 'application/json,text/plain,*/*',
            },
            timeout=timeout,
        )
        response.raise_for_status()
        payload = json.loads(response.content.decode('utf-8', errors='replace'))
        translated = ''.join(part[0] for part in payload[0] if part and part[0]).strip()
        if translated and translated.replace('?', '').strip():
            return translated


        # Logger for diagnostics
        logger = logging.getLogger(__name__)
    except Exception:
        pass

    try:
        translated = translate_text(text, target_lang=target_lang)
        if translated:
            return translated
    except Exception:
        pass

    return text


def format_mobile_number(value):
    """Format to '85758 69850' from phone-like input."""
    digits = re.sub(r'\D', '', str(value or ''))
    if len(digits) >= 10:
        digits = digits[-10:]
    return f"{digits[:5]} {digits[5:]}" if len(digits) == 10 else digits


def get_branch_bill_header_phones(branch):
    """Return formatted bill-header phone numbers from branch settings or fallback phone."""
    if not branch:
        return ''

    try:
        settings_obj = getattr(branch, 'settings', None)
    except (OperationalError, ProgrammingError):
        settings_obj = None
    raw_numbers = getattr(settings_obj, 'bill_header_mobile_numbers', '') if settings_obj else ''

    def _extract_numbers(text):
        extracted = []
        if not text:
            return extracted
        normalized = str(text).replace('\n', ',').replace('/', ',').replace('|', ',').replace(';', ',')
        for part in normalized.split(','):
            chunk = part.strip()
            if not chunk:
                continue
            digits = ''.join(ch for ch in chunk if ch.isdigit())
            if len(digits) >= 10:
                # Support pasted groups: pick each 10-digit slice from right.
                while len(digits) >= 10:
                    extracted.append(format_mobile_number(digits[-10:]))
                    digits = digits[:-10]
        return list(dict.fromkeys(extracted))

    numbers = _extract_numbers(raw_numbers)

    if numbers:
        return ', '.join(numbers)

    fallback_numbers = _extract_numbers(getattr(branch, 'phone', ''))
    if fallback_numbers:
        return ', '.join(fallback_numbers)
    return ''


def get_branch_bill_details(branch):
    """Return branch-wise bill header details with custom override support."""
    if not branch:
        return {
            'shop_name': 'Pawnshop Management System',
            'address': '',
            'phone': '',
            'email': '',
            'logo_url': '',
            'color_customer_name': '#000000',
            'color_issue_date': '#8B0000',
            'color_due_date': '#CC0000',
            'color_terms': '#111111',
            'color_items': '#000000',
            'color_header_title': '#002f6c',
            'color_header_subtitle': '#003f91',
            'color_header_text': '#444444',
            'color_field_label': '#333333',
            'color_field_value': '#000000',
            'color_table_header': '#222222',
        }

    default_address_parts = [branch.address, branch.city, branch.state, branch.zip_code]
    default_address = ', '.join([p for p in default_address_parts if p])

    details = {
        'shop_name': branch.name or 'Pawnshop Management System',
        'address': default_address,
        'phone': get_branch_bill_header_phones(branch),
        'email': branch.email or '',
        'logo_url': '',
        'color_customer_name': '#000000',
        'color_issue_date': '#8B0000',
        'color_due_date': '#CC0000',
        'color_terms': '#111111',
        'color_items': '#000000',
        'color_header_title': '#002f6c',
        'color_header_subtitle': '#003f91',
        'color_header_text': '#444444',
        'color_field_label': '#333333',
        'color_field_value': '#000000',
        'color_table_header': '#222222',
    }

    try:
        settings_obj = getattr(branch, 'settings', None)
    except (OperationalError, ProgrammingError):
        settings_obj = None

    if not settings_obj or not getattr(settings_obj, 'use_custom_bill_details', False):
        return details

    custom_shop_name = (getattr(settings_obj, 'bill_shop_name', '') or '').strip()
    custom_address = (getattr(settings_obj, 'bill_address', '') or '').strip()
    custom_email = (getattr(settings_obj, 'bill_email', '') or '').strip()
    custom_phone = (getattr(settings_obj, 'bill_header_mobile_numbers', '') or '').strip()

    if custom_shop_name:
        details['shop_name'] = custom_shop_name
    if custom_address:
        details['address'] = custom_address
    if custom_email:
        details['email'] = custom_email
    if custom_phone:
        numbers = []
        for part in custom_phone.replace('\n', ',').split(','):
            cleaned = part.strip()
            if cleaned:
                numbers.append(format_mobile_number(cleaned))
        if numbers:
            details['phone'] = ', '.join(numbers)

    logo = getattr(settings_obj, 'bill_logo', None)
    if logo:
        try:
            details['logo_url'] = logo.url
        except Exception:
            details['logo_url'] = ''

    details['color_customer_name'] = (getattr(settings_obj, 'bill_color_customer_name', '') or details['color_customer_name']).strip()
    details['color_issue_date'] = (getattr(settings_obj, 'bill_color_issue_date', '') or details['color_issue_date']).strip()
    details['color_due_date'] = (getattr(settings_obj, 'bill_color_due_date', '') or details['color_due_date']).strip()
    details['color_terms'] = (getattr(settings_obj, 'bill_color_terms', '') or details['color_terms']).strip()
    details['color_items'] = (getattr(settings_obj, 'bill_color_items', '') or details['color_items']).strip()
    details['color_header_title'] = (getattr(settings_obj, 'bill_color_header_title', '') or details['color_header_title']).strip()
    details['color_header_subtitle'] = (getattr(settings_obj, 'bill_color_header_subtitle', '') or details['color_header_subtitle']).strip()
    details['color_header_text'] = (getattr(settings_obj, 'bill_color_header_text', '') or details['color_header_text']).strip()
    details['color_field_label'] = (getattr(settings_obj, 'bill_color_field_label', '') or details['color_field_label']).strip()
    details['color_field_value'] = (getattr(settings_obj, 'bill_color_field_value', '') or details['color_field_value']).strip()
    details['color_table_header'] = (getattr(settings_obj, 'bill_color_table_header', '') or details['color_table_header']).strip()

    return details


def build_loan_pdf_language_context(loan, current_language):
    use_tamil = str(current_language).startswith('ta')
    customer = loan.customer
    loan_items = list(loan.loanitem_set.select_related('item'))
    branch = loan.branch

    customer_name_en = customer.full_name if customer else ''
    customer_name_ta = customer_name_en
    if customer and use_tamil:
        tamil_name_parts = [
            (customer.first_name_tamil or '').strip(),
            (customer.last_name_tamil or '').strip(),
        ]
        customer_name_ta = ' '.join([part for part in tamil_name_parts if part]).strip()

    customer_address_parts = []
    if customer:
        if customer.address:
            customer_address_parts.append(customer.address)
        if customer.city:
            customer_address_parts.append(customer.city)
        if customer.state:
            customer_address_parts.append(customer.state)
        if customer.zip_code:
            customer_address_parts.append(customer.zip_code)
    customer_address_en = ', '.join([part for part in customer_address_parts if part])
    customer_address_ta = customer_address_en
    if customer and use_tamil:
        tamil_address_parts = []
        if customer.address_tamil:
            tamil_address_parts.append(customer.address_tamil)
        if customer.city_tamil:
            tamil_address_parts.append(customer.city_tamil)
        if customer.state_tamil:
            tamil_address_parts.append(customer.state_tamil)
        if customer.zip_code:
            tamil_address_parts.append(customer.zip_code)
        customer_address_ta = ', '.join([part for part in tamil_address_parts if part])

    customer_id_label_en = customer.get_id_type_display() if customer and customer.id_type else ''
    customer_id_label_ta = customer_id_label_en

    branch_address_parts = []
    if branch:
        if branch.address:
            branch_address_parts.append(branch.address)
        if branch.city:
            branch_address_parts.append(branch.city)
        if branch.state:
            branch_address_parts.append(branch.state)
        if branch.zip_code:
            branch_address_parts.append(branch.zip_code)
    branch_address_en = ', '.join([part for part in branch_address_parts if part])
    branch_address_ta = branch_address_en
    customer_phone_display = format_mobile_number(customer.phone) if customer and getattr(customer, 'phone', None) else ''
    branch_phone_display = get_branch_bill_header_phones(branch)
    bill_details = get_branch_bill_details(branch)

    localized_items = []
    for loan_item in loan_items:
        item = loan_item.item
        name_en = item.name if item else ''
        description_en = item.description if item else ''
        name_ta = (item.tamil_name if item and getattr(item, 'tamil_name', '') else '') if use_tamil else ''
        description_ta = (item.tamil_description if item and getattr(item, 'tamil_description', '') else '') if use_tamil else ''

        localized_items.append({
            'loan_item': loan_item,
            'display_name': name_ta if use_tamil else name_en,
            'display_description': description_ta if use_tamil else description_en,
        })

    label_keys = {
        'document_title': 'Gold Loan Agreement',
        'borrower_name': 'Borrower Name',
        'loan_number': 'Loan Number',
        'email': 'Email',
        'phone': 'Phone Number',
        'address': 'Address',
        'id_details': 'ID Details',
        'not_provided': 'Not provided',
        'borrower_photo': 'Borrower Photo',
        'principal_amount': 'Principal Amount',
        'processing_fee': 'Processing Fee',
        'distribution_amount': 'Distribution Amount',
        'monthly_interest': 'Monthly Interest',
        'issue_date': 'Issue Date',
        'due_date': 'Due Date',
        'gold_items_details': 'Gold Items Details',
        'item_description': 'Item Description',
        'gold_karat': 'Gold Karat',
        'gross_weight': 'Gross Weight (g)',
        'net_weight': 'Net Weight (g)',
        'total_items': 'Total Items',
        'pledged_gold_item_photos': 'Pledged Gold Item Photos',
        'item': 'Item',
        'no_photos': 'No photos available for this loan.',
        'borrower_signature': 'Borrower Signature',
        'authorized_signatory': 'Authorized Signatory',
        'branch_manager': 'Branch Manager',
        'phone_label': 'Phone',
        'email_label': 'Email',
        'document_generated_on': 'Document generated on',
        'terms_and_conditions': 'TERMS AND CONDITIONS',
    }

    if use_tamil:
        labels = {
            'document_title': 'தங்கக் கடன் ஒப்பந்தம்',
            'borrower_name': 'கடன் வாங்கியவர் பெயர்',
            'loan_number': 'கடன் எண்',
            'email': 'மின்னஞ்சல்',
            'phone': 'தொலைபேசி எண்',
            'address': 'முகவரி',
            'id_details': 'அடையாள விவரங்கள்',
            'not_provided': 'வழங்கப்படவில்லை',
            'borrower_photo': 'கடன் வாங்கியவர் புகைப்படம்',
            'principal_amount': 'முதன்மை தொகை',
            'processing_fee': 'செயலாக்கக் கட்டணம்',
            'distribution_amount': 'வழங்கப்பட்ட தொகை',
            'monthly_interest': 'மாத வட்டி',
            'issue_date': 'வழங்கிய தேதி',
            'due_date': 'கடைசி தேதி',
            'gold_items_details': 'தங்கப் பொருட்கள் விவரம்',
            'item_description': 'பொருள் விவரம்',
            'gold_karat': 'தங்க சுத்தம்',
            'gross_weight': 'மொத்த எடை (கி)',
            'net_weight': 'நிகர எடை (கி)',
            'total_items': 'மொத்த பொருட்கள்',
            'pledged_gold_item_photos': 'அடமான பொருள் புகைப்படங்கள்',
            'item': 'பொருள்',
            'no_photos': 'இந்தக் கடனுக்கான படங்கள் எதுவுமில்லை.',
            'borrower_signature': 'கடன் வாங்கியவர் கையொப்பம்',
            'authorized_signatory': 'அங்கீகரிக்கப்பட்ட கையொப்பம்',
            'branch_manager': 'கிளை மேலாளர்',
            'phone_label': 'தொலைபேசி',
            'email_label': 'மின்னஞ்சல்',
            'document_generated_on': 'ஆவணம் உருவாக்கப்பட்ட தேதி',
            'terms_and_conditions': 'விதிமுறைகள் மற்றும் நிபந்தனைகள்',
        }
    else:
        labels = label_keys

    scheme_name = loan.scheme.name if loan.scheme else 'Standard Gold Loan'
    scheme_interest = loan.scheme.interest_rate if loan.scheme else loan.interest_rate
    scheme_duration = loan.scheme.loan_duration if loan.scheme and loan.scheme.loan_duration else 0
    display_scheme_name = scheme_name

    base_terms = [
        {
            'title': '1. Loan Scheme Details:',
            'content': f'This loan is issued under the "{scheme_name}" scheme. Interest rate: {scheme_interest}% per annum for {scheme_duration} days duration.'
        },
        {
            'title': '2. Purpose of Loan:',
            'content': 'The loan is granted solely on the security of gold ornaments/items deposited as collateral with the lender. The borrower affirms that the pledged article is their own property and is not stolen or encumbered.'
        },
        {
            'title': '3. Gold Recovery Timing:',
            'content': 'For gold recovery, payment must be made before 11:00 AM and gold collection will be available after 4:00 PM on the same day.'
        },
        {
            'title': '4. KYC Compliance:',
            'content': 'The borrower has provided necessary KYC documents as required under RBI guidelines, including proof of identity and address.'
        },
        {
            'title': '5. Fair Practices Code:',
            'content': "Loss or damage to the pledged article due to natural calamities, theft, or circumstances beyond the lender's control will not be the responsibility of the lender."
        },
        {
            'title': '6. Repayment and Recovery:',
            'content': 'The loan is repayable before the due date. If not repaid, the lender may sell the pledged article as per applicable law.'
        },
        {
            'title': '7. Receipt Requirement:',
            'content': 'No release of pledged gold items will be processed without verification of the original loan document and submitted ID proof.'
        },
        {
            'title': '8. Declaration:',
            'content': 'The borrower declares that all information provided is true and that they have read and understood all the terms and conditions mentioned herein.'
        },
    ]

    if use_tamil:
        terms = [
            {'title': '1. கடன் திட்ட விவரங்கள்:', 'content': f'இந்தக் கடன் "{display_scheme_name}" திட்டத்தின் கீழ் வழங்கப்படுகிறது. வட்டி விகிதம்: வருடத்திற்கு {scheme_interest}% மற்றும் காலம் {scheme_duration} நாட்கள்.'},
            {'title': '2. கடனின் நோக்கம்:', 'content': 'இந்தக் கடன் தங்கப் பொருட்களை அடமானமாக வைத்து வழங்கப்படுகிறது.'},
            {'title': '3. தங்கம் மீட்பு:', 'content': 'முதன்மை தொகையும் வட்டியும் முழுமையாக செலுத்திய பின் தங்கம் மீட்கப்படும்.'},
            {'title': '4. KYC இணக்கம்:', 'content': 'RBI வழிகாட்டுதலின்படி KYC விவரங்கள் சரிபார்க்கப்பட்டுள்ளன.'},
            {'title': '5. நியாய நடைமுறை:', 'content': 'கடன் வழங்கல் நடைமுறைகள் நியாயமாக பின்பற்றப்படும்.'},
            {'title': '6. திருப்பிச் செலுத்தல்:', 'content': 'கடைசி தேதிக்குள் கடன் தொகை செலுத்தப்பட வேண்டும்.'},
            {'title': '7. ரசீது அவசியம்:', 'content': 'ஒவ்வொரு கட்டணத்திற்கும் ரசீது வழங்கப்படும்.'},
            {'title': '8. அறிவிப்பு:', 'content': 'மேலே உள்ள தகவல்கள் அனைத்தும் உண்மை என கடன் வாங்கியவர் அறிவிக்கிறார்.'},
        ]
    else:
        terms = base_terms

    unique_item_names_count = len({
        (loan_item.item.name if loan_item.item and loan_item.item.name else '').strip().lower()
        for loan_item in loan_items
        if loan_item.item and loan_item.item.name
    })

    total_items_count = sum(
        int(getattr(loan_item, 'quantity', 1) or 1)
        for loan_item in loan_items
    ) if loan_items else 0
    if total_items_count <= 0:
        total_items_count = unique_item_names_count

    return {
        'current_language': current_language,
        'labels': labels,
        'localized_items': localized_items,
        'unique_item_names_count': unique_item_names_count,
        'total_items_count': total_items_count,
        'customer_name_display': customer_name_ta if use_tamil else customer_name_en,
        'customer_phone_display': customer_phone_display,
        'customer_address_display': customer_address_ta if use_tamil else customer_address_en,
        'customer_id_type_display': customer_id_label_ta if use_tamil else customer_id_label_en,
        'branch_address_display': branch_address_ta if use_tamil else branch_address_en,
        'branch_phone_display': branch_phone_display,
        'bill_shop_name': bill_details.get('shop_name', ''),
        'bill_address_display': bill_details.get('address', ''),
        'bill_phone_display': bill_details.get('phone', ''),
        'bill_email_display': bill_details.get('email', ''),
        'bill_logo_url': bill_details.get('logo_url', ''),
        'bill_color_customer_name': bill_details.get('color_customer_name', '#000000'),
        'bill_color_issue_date': bill_details.get('color_issue_date', '#8B0000'),
        'bill_color_due_date': bill_details.get('color_due_date', '#CC0000'),
        'bill_color_terms': bill_details.get('color_terms', '#111111'),
        'bill_color_items': bill_details.get('color_items', '#000000'),
        'bill_color_header_title': bill_details.get('color_header_title', '#002f6c'),
        'bill_color_header_subtitle': bill_details.get('color_header_subtitle', '#003f91'),
        'bill_color_header_text': bill_details.get('color_header_text', '#444444'),
        'bill_color_field_label': bill_details.get('color_field_label', '#333333'),
        'bill_color_field_value': bill_details.get('color_field_value', '#000000'),
        'bill_color_table_header': bill_details.get('color_table_header', '#222222'),
        'terms_list': terms,
    }


def get_loan_total_items_count(loan):
    """Return total item count for a loan, preferring LoanItem.quantity values."""
    if not loan:
        return 0
    loan_items = list(loan.loanitem_set.all())
    total = sum(int(getattr(item, 'quantity', 1) or 1) for item in loan_items) if loan_items else 0
    if total > 0:
        return total
    return len(loan_items)


def transliterate_between_english_tamil(request):
    text = (request.GET.get('text') or '').strip()
    direction = request.GET.get('direction', 'to_tamil')

    if not text:
        return JsonResponse({'result': ''})

    try:
        import requests

        source_lang = 'ta' if direction == 'to_english' else 'en'
        target_lang = 'en' if direction == 'to_english' else 'ta'
        translated_result = ''
        for _ in range(2):
            response = requests.get(
                'https://translate.googleapis.com/translate_a/single',
                params={
                    'client': 'gtx',
                    'sl': source_lang,
                    'tl': target_lang,
                    'dt': 't',
                    'q': text,
                },
                headers={
                    'User-Agent': 'Mozilla/5.0',
                    'Accept': 'application/json,text/plain,*/*',
                },
                timeout=10,
            )
            response.raise_for_status()
            payload = json.loads(response.content.decode('utf-8', errors='replace'))
            translated_result = ''.join(part[0] for part in payload[0] if part and part[0]).strip()
            if translated_result and translated_result != text and translated_result.replace('?', '').strip():
                return JsonResponse({'result': translated_result})
    except Exception:
        pass

    try:
        from indic_transliteration import sanscript
        from indic_transliteration.sanscript import transliterate

        if direction == 'to_english':
            result = transliterate(text, sanscript.TAMIL, sanscript.ITRANS)
        else:
            result = transliterate(text, sanscript.ITRANS, sanscript.TAMIL)

        return JsonResponse({'result': result})
    except Exception as exc:
        return JsonResponse({'result': text, 'error': str(exc)}, status=200)


# --- Number to words helpers (English and Tamil) ---
def amount_to_english_words(amount):
    """Convert a numeric amount to English words (rupees and paise).

    Falls back to simple formatting if num2words fails for a value.
    """
    try:
        dec = Decimal(amount)
    except Exception:
        return ''

    rupees = int(dec)
    paise = int((dec - rupees) * 100)

    try:
        # Prefer en_IN if available for Indian grouping, else en
        try:
            words_rupees = num2words(rupees, lang='en_IN')
        except Exception:
            words_rupees = num2words(rupees, lang='en')
        words = f"{words_rupees.capitalize()} rupees"
        if paise:
            words_paise = num2words(paise, lang='en')
            words += f" and {words_paise} paise"
        return words
    except Exception:
        # Fallback to simple string
        return f"{rupees} rupees{(' and ' + str(paise) + ' paise') if paise else ''}"


def _int_to_tamil_under_thousand(n):
    """Convert integer < 1000 to Tamil words."""
    ones = {
        0: 'ÃƒÂ Ã‚Â®Ã‚ÂªÃƒÂ Ã‚Â¯Ã¢â‚¬Å¡ÃƒÂ Ã‚Â®Ã…â€œÃƒÂ Ã‚Â¯Ã‚ÂÃƒÂ Ã‚Â®Ã‚Â¯ÃƒÂ Ã‚Â®Ã‚Â®ÃƒÂ Ã‚Â¯Ã‚Â', 1: 'ÃƒÂ Ã‚Â®Ã¢â‚¬â„¢ÃƒÂ Ã‚Â®Ã‚Â©ÃƒÂ Ã‚Â¯Ã‚ÂÃƒÂ Ã‚Â®Ã‚Â±ÃƒÂ Ã‚Â¯Ã‚Â', 2: 'ÃƒÂ Ã‚Â®Ã¢â‚¬Â¡ÃƒÂ Ã‚Â®Ã‚Â°ÃƒÂ Ã‚Â®Ã‚Â£ÃƒÂ Ã‚Â¯Ã‚ÂÃƒÂ Ã‚Â®Ã…Â¸ÃƒÂ Ã‚Â¯Ã‚Â', 3: 'ÃƒÂ Ã‚Â®Ã‚Â®ÃƒÂ Ã‚Â¯Ã¢â‚¬Å¡ÃƒÂ Ã‚Â®Ã‚Â©ÃƒÂ Ã‚Â¯Ã‚ÂÃƒÂ Ã‚Â®Ã‚Â±ÃƒÂ Ã‚Â¯Ã‚Â', 4: 'ÃƒÂ Ã‚Â®Ã‚Â¨ÃƒÂ Ã‚Â®Ã‚Â¾ÃƒÂ Ã‚Â®Ã‚Â©ÃƒÂ Ã‚Â¯Ã‚ÂÃƒÂ Ã‚Â®Ã¢â‚¬Â¢ÃƒÂ Ã‚Â¯Ã‚Â', 5: 'ÃƒÂ Ã‚Â®Ã‚ÂÃƒÂ Ã‚Â®Ã‚Â¨ÃƒÂ Ã‚Â¯Ã‚ÂÃƒÂ Ã‚Â®Ã‚Â¤ÃƒÂ Ã‚Â¯Ã‚Â',
        6: 'ÃƒÂ Ã‚Â®Ã¢â‚¬Â ÃƒÂ Ã‚Â®Ã‚Â±ÃƒÂ Ã‚Â¯Ã‚Â', 7: 'ÃƒÂ Ã‚Â®Ã‚ÂÃƒÂ Ã‚Â®Ã‚Â´ÃƒÂ Ã‚Â¯Ã‚Â', 8: 'ÃƒÂ Ã‚Â®Ã…Â½ÃƒÂ Ã‚Â®Ã…Â¸ÃƒÂ Ã‚Â¯Ã‚ÂÃƒÂ Ã‚Â®Ã…Â¸ÃƒÂ Ã‚Â¯Ã‚Â', 9: 'ÃƒÂ Ã‚Â®Ã¢â‚¬â„¢ÃƒÂ Ã‚Â®Ã‚Â©ÃƒÂ Ã‚Â¯Ã‚ÂÃƒÂ Ã‚Â®Ã‚ÂªÃƒÂ Ã‚Â®Ã‚Â¤ÃƒÂ Ã‚Â¯Ã‚Â', 10: 'ÃƒÂ Ã‚Â®Ã‚ÂªÃƒÂ Ã‚Â®Ã‚Â¤ÃƒÂ Ã‚Â¯Ã‚ÂÃƒÂ Ã‚Â®Ã‚Â¤ÃƒÂ Ã‚Â¯Ã‚Â', 11: 'ÃƒÂ Ã‚Â®Ã‚ÂªÃƒÂ Ã‚Â®Ã‚Â¤ÃƒÂ Ã‚Â®Ã‚Â¿ÃƒÂ Ã‚Â®Ã‚Â©ÃƒÂ Ã‚Â¯Ã…Â ÃƒÂ Ã‚Â®Ã‚Â©ÃƒÂ Ã‚Â¯Ã‚ÂÃƒÂ Ã‚Â®Ã‚Â±ÃƒÂ Ã‚Â¯Ã‚Â',
        12: 'ÃƒÂ Ã‚Â®Ã‚ÂªÃƒÂ Ã‚Â®Ã‚Â©ÃƒÂ Ã‚Â®Ã‚Â¿ÃƒÂ Ã‚Â®Ã‚Â°ÃƒÂ Ã‚Â®Ã‚Â£ÃƒÂ Ã‚Â¯Ã‚ÂÃƒÂ Ã‚Â®Ã…Â¸ÃƒÂ Ã‚Â¯Ã‚Â', 13: 'ÃƒÂ Ã‚Â®Ã‚ÂªÃƒÂ Ã‚Â®Ã‚Â¤ÃƒÂ Ã‚Â®Ã‚Â¿ÃƒÂ Ã‚Â®Ã‚Â©ÃƒÂ Ã‚Â¯Ã‚ÂÃƒÂ Ã‚Â®Ã‚Â®ÃƒÂ Ã‚Â¯Ã¢â‚¬Å¡ÃƒÂ Ã‚Â®Ã‚Â©ÃƒÂ Ã‚Â¯Ã‚ÂÃƒÂ Ã‚Â®Ã‚Â±ÃƒÂ Ã‚Â¯Ã‚Â', 14: 'ÃƒÂ Ã‚Â®Ã‚ÂªÃƒÂ Ã‚Â®Ã‚Â¤ÃƒÂ Ã‚Â®Ã‚Â¿ÃƒÂ Ã‚Â®Ã‚Â©ÃƒÂ Ã‚Â®Ã‚Â¾ÃƒÂ Ã‚Â®Ã‚Â©ÃƒÂ Ã‚Â¯Ã‚ÂÃƒÂ Ã‚Â®Ã¢â‚¬Â¢ÃƒÂ Ã‚Â¯Ã‚Â', 15: 'ÃƒÂ Ã‚Â®Ã‚ÂªÃƒÂ Ã‚Â®Ã‚Â¤ÃƒÂ Ã‚Â®Ã‚Â¿ÃƒÂ Ã‚Â®Ã‚Â©ÃƒÂ Ã‚Â¯Ã‹â€ ÃƒÂ Ã‚Â®Ã‚Â¨ÃƒÂ Ã‚Â¯Ã‚ÂÃƒÂ Ã‚Â®Ã‚Â¤ÃƒÂ Ã‚Â¯Ã‚Â',
        16: 'ÃƒÂ Ã‚Â®Ã‚ÂªÃƒÂ Ã‚Â®Ã‚Â¤ÃƒÂ Ã‚Â®Ã‚Â¿ÃƒÂ Ã‚Â®Ã‚Â©ÃƒÂ Ã‚Â®Ã‚Â¾ÃƒÂ Ã‚Â®Ã‚Â±ÃƒÂ Ã‚Â¯Ã‚Â', 17: 'ÃƒÂ Ã‚Â®Ã‚ÂªÃƒÂ Ã‚Â®Ã‚Â¤ÃƒÂ Ã‚Â®Ã‚Â¿ÃƒÂ Ã‚Â®Ã‚Â©ÃƒÂ Ã‚Â¯Ã¢â‚¬Â¡ÃƒÂ Ã‚Â®Ã‚Â´ÃƒÂ Ã‚Â¯Ã‚Â', 18: 'ÃƒÂ Ã‚Â®Ã‚ÂªÃƒÂ Ã‚Â®Ã‚Â¤ÃƒÂ Ã‚Â®Ã‚Â¿ÃƒÂ Ã‚Â®Ã‚Â©ÃƒÂ Ã‚Â¯Ã¢â‚¬Â ÃƒÂ Ã‚Â®Ã…Â¸ÃƒÂ Ã‚Â¯Ã‚ÂÃƒÂ Ã‚Â®Ã…Â¸ÃƒÂ Ã‚Â¯Ã‚Â', 19: 'ÃƒÂ Ã‚Â®Ã‚ÂªÃƒÂ Ã‚Â®Ã‚Â¤ÃƒÂ Ã‚Â¯Ã‚ÂÃƒÂ Ã‚Â®Ã‚Â¤ÃƒÂ Ã‚Â¯Ã…Â ÃƒÂ Ã‚Â®Ã‚Â©ÃƒÂ Ã‚Â¯Ã‚ÂÃƒÂ Ã‚Â®Ã‚ÂªÃƒÂ Ã‚Â®Ã‚Â¤ÃƒÂ Ã‚Â¯Ã‚Â'
    }
    tens = {
        20: 'ÃƒÂ Ã‚Â®Ã¢â‚¬Â¡ÃƒÂ Ã‚Â®Ã‚Â°ÃƒÂ Ã‚Â¯Ã‚ÂÃƒÂ Ã‚Â®Ã‚ÂªÃƒÂ Ã‚Â®Ã‚Â¤ÃƒÂ Ã‚Â¯Ã‚Â', 30: 'ÃƒÂ Ã‚Â®Ã‚Â®ÃƒÂ Ã‚Â¯Ã‚ÂÃƒÂ Ã‚Â®Ã‚ÂªÃƒÂ Ã‚Â¯Ã‚ÂÃƒÂ Ã‚Â®Ã‚ÂªÃƒÂ Ã‚Â®Ã‚Â¤ÃƒÂ Ã‚Â¯Ã‚Â', 40: 'ÃƒÂ Ã‚Â®Ã‚Â¨ÃƒÂ Ã‚Â®Ã‚Â¾ÃƒÂ Ã‚Â®Ã‚Â±ÃƒÂ Ã‚Â¯Ã‚ÂÃƒÂ Ã‚Â®Ã‚ÂªÃƒÂ Ã‚Â®Ã‚Â¤ÃƒÂ Ã‚Â¯Ã‚Â', 50: 'ÃƒÂ Ã‚Â®Ã‚ÂÃƒÂ Ã‚Â®Ã‚Â®ÃƒÂ Ã‚Â¯Ã‚ÂÃƒÂ Ã‚Â®Ã‚ÂªÃƒÂ Ã‚Â®Ã‚Â¤ÃƒÂ Ã‚Â¯Ã‚Â', 60: 'ÃƒÂ Ã‚Â®Ã¢â‚¬Â¦ÃƒÂ Ã‚Â®Ã‚Â±ÃƒÂ Ã‚Â¯Ã‚ÂÃƒÂ Ã‚Â®Ã‚ÂªÃƒÂ Ã‚Â®Ã‚Â¤ÃƒÂ Ã‚Â¯Ã‚Â',
        70: 'ÃƒÂ Ã‚Â®Ã…Â½ÃƒÂ Ã‚Â®Ã‚Â´ÃƒÂ Ã‚Â¯Ã‚ÂÃƒÂ Ã‚Â®Ã‚ÂªÃƒÂ Ã‚Â®Ã‚Â¤ÃƒÂ Ã‚Â¯Ã‚Â', 80: 'ÃƒÂ Ã‚Â®Ã…Â½ÃƒÂ Ã‚Â®Ã‚Â£ÃƒÂ Ã‚Â¯Ã‚ÂÃƒÂ Ã‚Â®Ã‚ÂªÃƒÂ Ã‚Â®Ã‚Â¤ÃƒÂ Ã‚Â¯Ã‚Â', 90: 'ÃƒÂ Ã‚Â®Ã‚Â¤ÃƒÂ Ã‚Â¯Ã…Â ÃƒÂ Ã‚Â®Ã‚Â£ÃƒÂ Ã‚Â¯Ã‚ÂÃƒÂ Ã‚Â®Ã‚Â£ÃƒÂ Ã‚Â¯Ã¢â‚¬Å¡ÃƒÂ Ã‚Â®Ã‚Â±ÃƒÂ Ã‚Â¯Ã‚Â'
    }

    parts = []
    if n >= 100:
        h = n // 100
        if h > 0:
            if h == 1:
                parts.append('ÃƒÂ Ã‚Â®Ã‚Â¨ÃƒÂ Ã‚Â¯Ã¢â‚¬Å¡ÃƒÂ Ã‚Â®Ã‚Â±ÃƒÂ Ã‚Â¯Ã‚Â')
            else:
                parts.append(ones.get(h, '') + ' ÃƒÂ Ã‚Â®Ã‚Â¨ÃƒÂ Ã‚Â¯Ã¢â‚¬Å¡ÃƒÂ Ã‚Â®Ã‚Â±ÃƒÂ Ã‚Â¯Ã‚Â')
        n = n % 100

    if n >= 20:
        t = (n // 10) * 10
        if t in tens:
            parts.append(tens[t])
            r = n % 10
            if r:
                parts.append(ones.get(r, ''))
        else:
            parts.append(ones.get(n, ''))
    elif n > 0:
        parts.append(ones.get(n, ''))

    return ' '.join([p for p in parts if p])


def number_to_tamil_words(amount):
    """Convert numeric amount to Tamil words (Indian grouping)."""
    try:
        dec = Decimal(amount)
    except Exception:
        return ''

    rupees = int(dec)
    paise = int((dec - rupees) * 100)

    if rupees == 0:
        rupees_words = 'ÃƒÂ Ã‚Â®Ã…Â¡ÃƒÂ Ã‚Â¯Ã‚ÂÃƒÂ Ã‚Â®Ã‚Â´ÃƒÂ Ã‚Â®Ã‚Â¿ÃƒÂ Ã‚Â®Ã‚Â¯ÃƒÂ Ã‚Â®Ã‚Â®ÃƒÂ Ã‚Â¯Ã‚Â'
    else:
        parts = []
        crore = rupees // 10000000
        if crore:
            parts.append(f"{_int_to_tamil_under_thousand(crore)} ÃƒÂ Ã‚Â®Ã¢â‚¬Â¢ÃƒÂ Ã‚Â¯Ã¢â‚¬Â¹ÃƒÂ Ã‚Â®Ã…Â¸ÃƒÂ Ã‚Â®Ã‚Â¿")
        rupees = rupees % 10000000

        lakh = rupees // 100000
        if lakh:
            parts.append(f"{_int_to_tamil_under_thousand(lakh)} ÃƒÂ Ã‚Â®Ã‚Â²ÃƒÂ Ã‚Â®Ã…Â¸ÃƒÂ Ã‚Â¯Ã‚ÂÃƒÂ Ã‚Â®Ã…Â¡ÃƒÂ Ã‚Â®Ã‚Â®ÃƒÂ Ã‚Â¯Ã‚Â")
        rupees = rupees % 100000

        thousand = rupees // 1000
        if thousand:
            parts.append(f"{_int_to_tamil_under_thousand(thousand)} ÃƒÂ Ã‚Â®Ã¢â‚¬Â ÃƒÂ Ã‚Â®Ã‚Â¯ÃƒÂ Ã‚Â®Ã‚Â¿ÃƒÂ Ã‚Â®Ã‚Â°ÃƒÂ Ã‚Â®Ã‚Â®ÃƒÂ Ã‚Â¯Ã‚Â")
        rupees = rupees % 1000

        if rupees:
            parts.append(_int_to_tamil_under_thousand(rupees))

        rupees_words = ' '.join([p for p in parts if p])

    result = f"{rupees_words} ÃƒÂ Ã‚Â®Ã‚Â°ÃƒÂ Ã‚Â¯Ã¢â‚¬Å¡ÃƒÂ Ã‚Â®Ã‚ÂªÃƒÂ Ã‚Â®Ã‚Â¾ÃƒÂ Ã‚Â®Ã‚Â¯ÃƒÂ Ã‚Â¯Ã‚Â"
    if paise:
        paise_words = _int_to_tamil_under_thousand(paise)
        result = f"{result} {paise_words} ÃƒÂ Ã‚Â®Ã‚ÂªÃƒÂ Ã‚Â®Ã‚Â£ÃƒÂ Ã‚Â¯Ã‹â€ "
    return result


def process_item_photos_for_display(item_photos):
    """
    Centralized function to process item photos for display across the project.
    Handles both old file-based photos and new database-stored base64 photos.
    
    Args:
        item_photos: String containing either JSON array of photos or single photo data
        
    Returns:
        list: List of photo URLs/data URLs ready for display
    """
    if not item_photos:
        return []
    
    try:
        photo_list = []
        
        # Handle single base64 image
        if isinstance(item_photos, str) and item_photos.startswith('data:image/'):
            return [item_photos]
        
        # Handle JSON array of photos
        if isinstance(item_photos, str):
            if item_photos.startswith('['):
                photos_data = json.loads(item_photos)
            else:
                photos_data = [item_photos]
        else:
            photos_data = item_photos if isinstance(item_photos, list) else [item_photos]
        
        for photo in photos_data:
            if photo and isinstance(photo, str):
                if photo.startswith('data:image/'):
                    # Already base64 format, use directly
                    photo_list.append(photo)
                elif photo.startswith('/media/'):
                    # File path - convert to base64 or check if file exists
                    try:
                        relative_path = photo.replace('/media/', '')
                        file_path = os.path.join(settings.MEDIA_ROOT, relative_path)
                        
                        if os.path.exists(file_path):
                            # Convert file to base64 for consistent display
                            with open(file_path, 'rb') as f:
                                file_content = f.read()
                                encoded = base64.b64encode(file_content).decode('utf-8')
                                photo_list.append(f"data:image/jpeg;base64,{encoded}")
                        else:
                            # File doesn't exist, skip it
                            continue
                    except Exception as e:
                        print(f"Error processing file photo {photo}: {str(e)}")
                        continue
                else:
                    # Assume it's already base64 (without data: prefix)
                    photo_list.append(f"data:image/jpeg;base64,{photo}")
        
        return photo_list
    
    except Exception as e:
        print(f"Error processing item photos: {str(e)}")
        return []


def get_first_item_photo(item_photos):
    """
    Get the first item photo for thumbnails and previews.
    
    Args:
        item_photos: String containing photo data
        
    Returns:
        str: First photo URL/data URL or placeholder if none available
    """
    photos = process_item_photos_for_display(item_photos)
    if photos:
        return photos[0]
    return "/static/img/placeholder-item.png"


def get_item_photos_count(item_photos):
    """
    Get the count of item photos.
    
    Args:
        item_photos: String containing photo data
        
    Returns:
        int: Number of photos
    """
    photos = process_item_photos_for_display(item_photos)
    return len(photos)


# Basic placeholder views for the transactions app
# These will need to be implemented properly with the correct models

class LoanListView(LoginRequiredMixin, RoleBranchAccessMixin, DownloadMixin, ListView):
    model = Loan
    template_name = 'transactions/loan_list.html'
    context_object_name = 'loans'
    paginate_by = 15
    
    def get_queryset(self):
        queryset = Loan.objects.all()
        user = self.request.user

        # Apply centralized branch/region access rules and organization isolation
        queryset = self.filter_queryset_by_branches(queryset, branch_field_name='branch')
        if user.organization:
            queryset = queryset.filter(branch__organization=user.organization)

        # Status filter
        status = self.request.GET.get('status')
        if status:
            queryset = queryset.filter(status=status)

        # Search filter
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(customer__first_name__icontains=search) |
                Q(customer__last_name__icontains=search) |
                Q(loan_number__icontains=search) |
                Q(loanitem__item__name__icontains=search)
            ).distinct()

        # Date range filter
        date_range = self.request.GET.get('date_range')
        today = timezone.now().date()
        
        if date_range == 'today':
            queryset = queryset.filter(issue_date=today)
        elif date_range == 'this_week':
            week_start = today - timezone.timedelta(days=today.weekday())
            queryset = queryset.filter(issue_date__gte=week_start)
        elif date_range == 'this_month':
            queryset = queryset.filter(issue_date__year=today.year, issue_date__month=today.month)
        elif date_range == 'this_year':
            queryset = queryset.filter(issue_date__year=today.year)

        # New filter for overdue loans
        filter_type = self.request.GET.get('filter_type')
        if filter_type == 'overdue':
            queryset = queryset.filter(status='active', due_date__lt=today)
        elif filter_type == 'due_soon':
            queryset = queryset.filter(status='active', due_date__gte=today, due_date__lte=today + timezone.timedelta(days=30))

        # Sorting
        sort_by = self.request.GET.get('sort', '-issue_date')  # Default sort by newest first
        valid_sort_fields = {
            'customer': 'customer__first_name',
            '-customer': '-customer__first_name',
            'principal': 'principal_amount',
            '-principal': '-principal_amount',
            'issue_date': 'issue_date',
            '-issue_date': '-issue_date',
            'due_date': 'due_date',
            '-due_date': '-due_date',
            'status': 'status',
            '-status': '-status',
            'branch': 'branch__name',
            '-branch': '-branch__name',
        }
        
        if sort_by in valid_sort_fields:
            queryset = queryset.order_by(valid_sort_fields[sort_by])
        else:
            queryset = queryset.order_by('-issue_date')  # Default fallback

        return queryset.select_related('customer', 'branch').prefetch_related('loanitem_set', 'loanitem_set__item')

    def get_download_filename(self, format_type='csv'):
        """Generate download filename for loans export"""
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        status = (self.request.GET.get('status') or 'all').strip().lower()
        # Normalize status to safe filename fragment
        status = re.sub(r'[^a-z0-9_\-]', '_', status)
        return f'loans_export_{status}_{timestamp}.{format_type}'

    def get_download_headers(self):
        """Return headers for download export"""
        all_headers = [
            ('roll_number', 'Roll Number'),
            ('loan_number', 'Loan Number'),
            ('customer_name', 'Customer Name'),
            ('phone', 'Customer Phone'),
            ('email', 'Customer Email'), 
            ('branch', 'Branch'),
            ('item_images', 'Image Count'),
            ('principal_amount', 'Principal Amount (₹)'),
            ('distribution_amount', 'Distribution Amount (₹)'),
            ('interest_rate', 'Interest Rate (%)'),
            ('issue_date', 'Issue Date'),
            ('due_date', 'Due Date'),
            ('status', 'Status'),
            ('days_since_issue', 'Days Since Issue'),
            ('days_remaining', 'Days Remaining'),
            ('item_names', 'Item Names'),
            ('total_weight', 'Total Weight (grams)'),
            ('karat', 'Gold Karat'),
            ('monthly_interest', 'Monthly Interest Amount (₹)'),
            ('total_payable', 'Total Payable Till Date (₹)'),
            ('amount_paid', 'Amount Paid (₹)'),
            ('remaining_balance', 'Remaining Balance (₹)'),
            ('created_at', 'Created Date'),
            ('created_by', 'Created By')
        ]
        
        selected_columns = self.get_selected_columns()
        if selected_columns:
            headers = []
            for col_key, col_name in all_headers:
                if col_key in selected_columns:
                    headers.append(col_name)
            return headers
        
        return [col_name for _, col_name in all_headers]
    
    def get_selected_columns(self):
        """Get selected columns from request"""
        columns_param = self.request.GET.get('columns', '')
        if columns_param:
            return columns_param.split(',')
        return None
    
    def filter_row_data(self, row_data, selected_columns=None):
        """Filter row data based on selected columns"""
        if not selected_columns:
            return row_data
        
        column_keys = [
            'roll_number', 'loan_number', 'customer_name', 'phone', 'email', 'branch',
            'item_images', 'principal_amount', 'distribution_amount', 'interest_rate', 'issue_date', 'due_date', 'status',
            'days_since_issue', 'days_remaining', 'item_names', 'total_weight', 'karat',
            'monthly_interest', 'total_payable', 'amount_paid', 'remaining_balance',
            'created_at', 'created_by'
        ]
        
        filtered_row = []
        for i, col_key in enumerate(column_keys):
            if i < len(row_data) and col_key in selected_columns:
                filtered_row.append(row_data[i])
        return filtered_row

    def get_download_data(self):
        """Return data for download export"""
        queryset = self.get_queryset()
        
        data = []
        for index, loan in enumerate(queryset, start=1):
            # Get loan items information
            loan_items = loan.loanitem_set.all()
            item_names = []
            total_weight = 0
            karat_info = set()
            
            for item in loan_items:
                if item.item:
                    item_names.append(item.item.name)
                if hasattr(item, 'net_weight') and item.net_weight:
                    total_weight += float(item.net_weight)
                if hasattr(item, 'gold_karat') and item.gold_karat:
                    karat_info.add(f"{item.gold_karat}K")
            
            # Calculate financial information
            try:
                monthly_interest = 0
                if hasattr(loan, 'monthly_interest_amount'):
                    monthly_interest = float(loan.monthly_interest_amount())
                
                total_payable = 0
                if hasattr(loan, 'total_payable_till_date'):
                    total_payable = float(loan.total_payable_till_date)
                
                amount_paid = 0
                if hasattr(loan, 'amount_paid'):
                    amount_paid = float(loan.amount_paid)
                
                remaining_balance = total_payable - amount_paid
            except:
                monthly_interest = 0
                total_payable = 0
                amount_paid = 0
                remaining_balance = 0
            
            # Calculate days information
            try:
                days_since_issue = (timezone.now().date() - loan.issue_date).days if loan.issue_date else 0
                days_remaining = (loan.due_date - timezone.now().date()).days if loan.due_date else 0
            except:
                days_since_issue = 0
                days_remaining = 0
            
            row = [
                index,
                loan.loan_number or '',
                f"{loan.customer.first_name} {loan.customer.last_name}" if loan.customer else '',
                loan.customer.phone if loan.customer and hasattr(loan.customer, 'phone') else '',
                loan.customer.email if loan.customer and hasattr(loan.customer, 'email') else '',
                loan.branch.name if loan.branch else '',
                str(len(loan.item_photo_list)) if hasattr(loan, 'item_photo_list') and loan.item_photo_list else '0',
                float(loan.principal_amount) if loan.principal_amount else 0,
                float(loan.distribution_amount) if hasattr(loan, 'distribution_amount') and loan.distribution_amount else 0,
                float(loan.interest_rate) if loan.interest_rate else 0,
                loan.issue_date.strftime('%Y-%m-%d') if loan.issue_date else '',
                loan.due_date.strftime('%Y-%m-%d') if loan.due_date else '',
                loan.get_status_display() if hasattr(loan, 'get_status_display') else (loan.status or ''),
                days_since_issue,
                days_remaining,
                ', '.join(item_names) if item_names else '',
                total_weight,
                ', '.join(sorted(karat_info)) if karat_info else '',
                monthly_interest,
                total_payable,
                amount_paid,
                remaining_balance,
                loan.created_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(loan, 'created_at') and loan.created_at else '',
                f"{loan.created_by.first_name} {loan.created_by.last_name}" if hasattr(loan, 'created_by') and loan.created_by else ''
            ]
            data.append(row)
        
        return data

    def get_download_data_for_index(self, loan, index):
        """Return data row for a single loan (used by PDF export)."""
        try:
            monthly_interest = 0
            if hasattr(loan, 'monthly_interest_amount'):
                monthly_interest = float(loan.monthly_interest_amount())

            total_payable = 0
            if hasattr(loan, 'total_payable_till_date'):
                total_payable = float(loan.total_payable_till_date)

            amount_paid = 0
            if hasattr(loan, 'amount_paid'):
                amount_paid = float(loan.amount_paid)

            remaining_balance = total_payable - amount_paid
        except Exception:
            monthly_interest = 0
            total_payable = 0
            amount_paid = 0
            remaining_balance = 0

        try:
            days_since_issue = (timezone.now().date() - loan.issue_date).days if loan.issue_date else 0
            days_remaining = (loan.due_date - timezone.now().date()).days if loan.due_date else 0
        except Exception:
            days_since_issue = 0
            days_remaining = 0

        loan_items = loan.loanitem_set.all()
        item_names = [li.item.name for li in loan_items if li.item]

        return [
            index,
            loan.loan_number or '',
            f"{loan.customer.first_name} {loan.customer.last_name}" if loan.customer else '',
            loan.customer.phone if loan.customer and hasattr(loan.customer, 'phone') else '',
            loan.customer.email if loan.customer and hasattr(loan.customer, 'email') else '',
            loan.branch.name if loan.branch else '',
            str(len(getattr(loan, 'item_photo_list', []))) if getattr(loan, 'item_photo_list', None) else '0',
            float(loan.principal_amount) if loan.principal_amount else 0,
            float(loan.distribution_amount) if hasattr(loan, 'distribution_amount') and loan.distribution_amount else 0,
            float(loan.interest_rate) if loan.interest_rate else 0,
            loan.issue_date.strftime('%Y-%m-%d') if loan.issue_date else '',
            loan.due_date.strftime('%Y-%m-%d') if loan.due_date else '',
            loan.get_status_display() if hasattr(loan, 'get_status_display') else (loan.status or ''),
            days_since_issue,
            days_remaining,
            ', '.join(item_names) if item_names else '',
            0,
            '',
            monthly_interest,
            total_payable,
            amount_paid,
            remaining_balance,
            loan.created_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(loan, 'created_at') and loan.created_at else '',
            f"{loan.created_by.first_name} {loan.created_by.last_name}" if hasattr(loan, 'created_by') and loan.created_by else ''
        ]

    
    def download_csv(self):
        return self.export_csv()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        context['selected_status'] = self.request.GET.get('status', '')
        context['selected_date_range'] = self.request.GET.get('date_range', '')
        context['selected_filter_type'] = self.request.GET.get('filter_type', '')
        context['current_sort'] = self.request.GET.get('sort', '-issue_date')
        
        # Calculate loan statistics for the cards - optimized with aggregation
        user = self.request.user
        base_queryset = Loan.objects.all()
        base_queryset = self.filter_queryset_by_branches(base_queryset, branch_field_name='branch')
        if user.organization:
            base_queryset = base_queryset.filter(branch__organization=user.organization)
        
        # Calculate statistics efficiently - SINGLE QUERY with annotations
        from django.utils import timezone
        from django.db.models import Count, Q, Sum, F, Case, When
        from decimal import Decimal
        
        today = timezone.now().date()
        
        # Get all statistics in a single aggregation query, including outstanding sums
        from django.db.models import Sum, F, Value as V
        from django.db.models.functions import Coalesce

        # Expression for outstanding per loan: use total_payable_till_date if present else principal_amount, minus amount_paid
        outstanding_expr = (Coalesce(F('total_payable_till_date'), F('principal_amount'), V(0)) - Coalesce(F('amount_paid'), V(0)))

        # Only aggregate simple count metrics via the database; monetary
        # totals rely on model properties and are calculated in Python below.
        stats = base_queryset.aggregate(
            active_count=Count('id', filter=Q(status='active')),
            due_today_count=Count('id', filter=Q(status='active', due_date=today)),
            overdue_count=Count('id', filter=Q(status='active', due_date__lt=today)),
            due_soon_count=Count('id', filter=Q(status='active', due_date__gte=today, due_date__lte=today + timezone.timedelta(days=30))),
        )

        # Debug logging to help trace incorrect zeros in the UI
        try:
            import logging
            logger = logging.getLogger('transactions.views')
            logger.info(f"LoanListView.get_context_data called for user={getattr(user,'username',None)}; stats={stats}")
        except Exception:
            pass
        # Also print to stdout for dev server visibility
        try:
            print(f"[DEBUG] LoanListView.stats for user={getattr(user,'username',None)}: {stats}")
        except Exception:
            pass

        context['active_loans_count'] = stats.get('active_count') or 0
        context['due_today_count'] = stats.get('due_today_count') or 0
        context['overdue_count'] = stats.get('overdue_count') or 0
        context['due_soon_count'] = stats.get('due_soon_count') or 0

        # Monetary summaries (Decimal) - coerce None to 0
        # Compute monetary summaries in Python using model properties (accurate
        # even when values are computed via methods). Prefetch payments to avoid
        # N+1 queries.
        loans_iter = base_queryset.select_related('customer', 'branch').prefetch_related('payments')

        def loan_outstanding(ln):
            try:
                tp = getattr(ln, 'total_payable_till_date', None)
                if callable(tp):
                    tp = tp()
                if tp is None:
                    tp = getattr(ln, 'principal_amount', 0) or 0

                ap = getattr(ln, 'amount_paid', None)
                if callable(ap):
                    ap = ap()
                if ap is None:
                    try:
                        ap = sum(p.amount for p in (ln.payments.all() if hasattr(ln, 'payments') else []))
                    except Exception:
                        ap = 0

                out = Decimal(tp or 0) - Decimal(ap or 0)
                return out if out > 0 else Decimal('0.00')
            except Exception:
                return Decimal('0.00')

        active_sum = Decimal('0.00')
        due_today_sum = Decimal('0.00')
        due_soon_sum = Decimal('0.00')
        overdue_sum = Decimal('0.00')
        total_sum = Decimal('0.00')
        
        thirty_days_later = today + timezone.timedelta(days=30)

        for ln in loans_iter:
            o = loan_outstanding(ln)
            total_sum += o
            if getattr(ln, 'status', '') == 'active':
                active_sum += o
                if getattr(ln, 'due_date', None) == today:
                    due_today_sum += o
                if getattr(ln, 'due_date', None) and getattr(ln, 'due_date') < today:
                    overdue_sum += o
                if getattr(ln, 'due_date', None) and today <= getattr(ln, 'due_date') <= thirty_days_later:
                    due_soon_sum += o

        context['active_outstanding'] = active_sum
        context['due_today_outstanding'] = due_today_sum
        context['due_soon_outstanding'] = due_soon_sum
        context['overdue_outstanding'] = overdue_sum
        context['total_outstanding'] = total_sum

        # Visibility controls: show financial summary only for admin users
        is_admin_user = bool((user.username == 'admin') or user.is_staff or user.is_superuser or getattr(user, 'is_pawnshop_admin', False) or getattr(user, 'is_organization_admin', False))
        context['show_total_outstanding'] = is_admin_user
        context['show_total_loan_lists'] = is_admin_user
        return context

    def get(self, request, *args, **kwargs):
        # Check if download is requested
        download_format = request.GET.get('download')
        if download_format:
            if download_format == 'csv':
                return self.export_csv()
            elif download_format == 'excel':
                return self.export_excel()
            elif download_format == 'pdf':
                return self.export_pdf()
        
        return super().get(request, *args, **kwargs)

    def export_csv(self):
        """Export data as CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{self.get_download_filename("csv")}"'
        
        writer = csv.writer(response)
        writer.writerow(self.get_download_headers())
        
        selected_columns = self.get_selected_columns()
        for row in self.get_download_data():
            filtered_row = self.filter_row_data(row, selected_columns)
            writer.writerow(filtered_row)
        
        return response

    def export_excel(self):
        """Export data as Excel"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Loans Export"
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        headers = self.get_download_headers()
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        selected_columns = self.get_selected_columns()
        for row_idx, row_data in enumerate(self.get_download_data(), 2):
            filtered_row = self.filter_row_data(row_data, selected_columns)
            for col_idx, value in enumerate(filtered_row, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to response
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.get_download_filename("xlsx")}"'
        
        return response

    def export_pdf(self):
        from reportlab.lib.pagesizes import letter, A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from datetime import datetime
        
        response = HttpResponse(content_type='application/pdf')
        disposition = 'inline' if self.request.GET.get('preview') == '1' else 'attachment'
        response['Content-Disposition'] = f'{disposition}; filename="{self.get_download_filename("pdf")}"'
        
        doc = SimpleDocTemplate(
            response,
            pagesize=landscape(A4),
            rightMargin=0.3*inch,
            leftMargin=0.3*inch,
            topMargin=0.6*inch,
            bottomMargin=0.6*inch
        )
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=14,
            textColor=colors.HexColor('#2C3E50'),
            spaceAfter=12,
            alignment=1
        )
        elements.append(Paragraph("LOANS EXPORT", title_style))
        elements.append(Spacer(1, 12))

        header_cell_style = ParagraphStyle(
            'LoanExportHeaderCell',
            parent=styles['Normal'],
            fontName='Helvetica-Bold',
            fontSize=7.5,
            leading=9,
            textColor=colors.whitesmoke,
            alignment=TA_CENTER,
            splitLongWords=1,
        )
        body_cell_style = ParagraphStyle(
            'LoanExportBodyCell',
            parent=styles['Normal'],
            fontName='Helvetica',
            fontSize=7,
            leading=8.5,
            alignment=TA_LEFT,
            splitLongWords=1,
        )
        
        # PDF Export - Show only 10 key columns for clarity
        pdf_headers = [
            'Roll Number',
            'Loan Number',
            'Customer Name',
            'Customer Phone',
            'Branch',
            'Distribution Amount (₹)',
            'Issue Date',
            'Due Date',
            'Status',
            'Item Names'
        ]
        
        # Map header names to row indices
        all_headers = self.get_download_headers()
        header_indices = {
            'Roll Number': 0,
            'Loan Number': 1,
            'Customer Name': 2,
            'Customer Phone': 3,
            'Branch': 5,
            'Distribution Amount (₹)': 8,
            'Issue Date': 10,
            'Due Date': 11,
            'Status': 12,
            'Item Names': 15
        }
        
        if self.request.GET.get('full') == '1':
            pdf_headers = self.get_download_headers()
            header_indices = {header: index for index, header in enumerate(pdf_headers)}

        # Build table with only selected columns
        table_data = [[Paragraph(str(header), header_cell_style) for header in pdf_headers]]
        queryset = self.get_queryset()
        
        for index, loan in enumerate(queryset, start=1):
            row = self.get_download_data_for_index(loan, index)
            pdf_row = [
                Paragraph(str(row[header_indices[h]]) if row[header_indices[h]] is not None else '', body_cell_style)
                for h in pdf_headers
            ]
            table_data.append(pdf_row)
        
        # Column widths for 10 columns on landscape A4
        # Available width: 11.69 - 0.6 = 11.09 inches
        col_widths = [
            0.65 * inch,  # Roll Number
            1.05 * inch,  # Loan Number
            1.25 * inch,  # Customer Name
            0.95 * inch,  # Customer Phone
            1.35 * inch,  # Branch
            0.95 * inch,  # Distribution Amount
            0.8 * inch,   # Issue Date
            0.8 * inch,   # Due Date
            0.65 * inch,  # Status
            2.0 * inch,   # Item Names
        ]
        
        if self.request.GET.get('full') == '1':
            col_widths = [0.45 * inch] * len(pdf_headers)

        # Create table with specified column widths
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        
        # Define table style with better formatting
        table_style = TableStyle([
            # Header style - blue background with white text
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, 0), 5),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 5),
            
            # Data rows - larger font for readability
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('VALIGN', (0, 1), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ('RIGHTPADDING', (0, 0), (-1, -1), 3),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            
            # Grid lines - darker and thicker for visibility
            ('GRID', (0, 0), (-1, -1), 1.5, colors.HexColor('#34495E')),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#ECF0F1')]),
            
            # Right align numeric columns
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Roll Number
            ('ALIGN', (5, 1), (5, -1), 'RIGHT'),   # Distribution Amount
            ('ALIGN', (6, 1), (8, -1), 'CENTER'),  # Dates, Status
        ])
        
        table.setStyle(table_style)
        elements.append(table)
        elements.append(Spacer(1, 12))
        
        # Footer with timestamp
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.HexColor('#7F8C8D'),
            alignment=1
        )
        elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%d %B %Y at %I:%M %p')}", footer_style))
        
        # Build PDF
        doc.build(elements)
        return response


class LoanExpiryNoticeView(LoginRequiredMixin, RoleBranchAccessMixin, View):
    def get(self, request, loan_number):
        loan = get_object_or_404(Loan, loan_number=loan_number)
        # enforce branch access
        self.check_object_branch_access(loan, branch_attr='branch')
        today = timezone.now().date()

        # Determine if an expiry/auction notice should be shown
        show_notice = False
        try:
            if getattr(loan, 'due_date', None) and loan.due_date <= today:
                show_notice = True
            elif getattr(loan, 'grace_period_end', None) and loan.grace_period_end <= today:
                show_notice = True
        except Exception:
            show_notice = False

        # Compute a best-effort remaining balance for display
        remaining = None
        try:
            remaining = getattr(loan, 'remaining_balance', None)
            if remaining is None:
                total_payable = getattr(loan, 'total_payable_till_date', None) or 0
                paid = sum(p.amount for p in loan.payments.all()) if hasattr(loan, 'payments') else 0
                remaining = max(0, total_payable - paid) if total_payable else getattr(loan, 'principal_amount', 0)
        except Exception:
            remaining = getattr(loan, 'principal_amount', 0)

        use_tamil = str(getattr(request, 'LANGUAGE_CODE', '')).startswith('ta')

        context = {
            'loan': loan,
            'remaining_balance': remaining,
            'today': today,
            'show_notice': show_notice,
            'use_tamil': use_tamil,
            'branch_phone_display': get_branch_bill_header_phones(getattr(loan, 'branch', None)),
        }
        return render(request, 'transactions/loan_expiry_notice.html', context)

class LoanDetailView(LoginRequiredMixin, RoleBranchAccessMixin, DetailView):
    model = Loan
    template_name = 'transactions/loan_detail.html'
    context_object_name = 'loan'
    slug_field = 'loan_number'
    slug_url_kwarg = 'loan_number'

    def get_object(self, queryset=None):
        obj = super().get_object(queryset=queryset)
        # enforce branch access
        self.check_object_branch_access(obj, branch_attr='branch')
        return obj

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        loan = self.get_object()
        
        # Add payments and other related data to context
        context['payments'] = loan.payments.all().order_by('-payment_date')
        context['extensions'] = loan.extensions.all().order_by('-extension_date')
        context['loan_items'] = loan.loanitem_set.all()
        
        # Process item photos for the template using centralized function
        context['item_photos_list'] = process_item_photos_for_display(loan.item_photos)
        print(f"Retrieved {len(context['item_photos_list'])} item photos from database for loan {loan.loan_number}")
            
        return context


class LoanCreateView(LoginRequiredMixin, RoleBranchAccessMixin, CreateView):
    model = Loan
    form_class = LoanForm
    template_name = 'transactions/loan_form.html'
    
    def get_success_url(self):
        return reverse('loan_detail', kwargs={'loan_number': self.object.loan_number})
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def get_initial(self):
        """Set initial values for the form, including pre-selected customer"""
        initial = super().get_initial()
        
        # Handle customer_id from URL parameter (from customer detail/list pages)
        customer_id = self.request.GET.get('customer_id')
        if customer_id:
            try:
                from accounts.models import Customer
                customer = Customer.objects.get(id=customer_id)
                # Verify the customer belongs to the user's organization/branch
                user = self.request.user
                if user.organization:
                    if customer.branch and customer.branch.organization == user.organization:
                        initial['customer'] = customer
                        # Also set the branch to customer's branch if available
                        if customer.branch:
                            initial['branch'] = customer.branch
                elif user.is_superuser:
                    # Superusers can access any customer
                    initial['customer'] = customer
                    if customer.branch:
                        initial['branch'] = customer.branch
            except Customer.DoesNotExist:
                # Customer not found, ignore the parameter
                pass
        
        return initial
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Add customer information to context if pre-selected
        customer_id = self.request.GET.get('customer_id')
        if customer_id:
            try:
                from accounts.models import Customer
                customer = Customer.objects.get(id=customer_id)
                context['preselected_customer'] = customer
                context['customer_preselected'] = True
            except Customer.DoesNotExist:
                pass
        
        # Process item photos for form if editing existing loan
        if self.object and self.object.item_photos:
            context['item_photos_list'] = process_item_photos_for_display(self.object.item_photos)
        return context
    
    def form_valid(self, form):
        # Handle photo processing during loan creation
        item_photos_data = self.request.POST.get('item_photos', '')
        customer_face_capture = self.request.POST.get('customer_face_capture', '')
        
        # Save photos directly to database
        if item_photos_data:
            form.instance.item_photos = item_photos_data
        if customer_face_capture:
            form.instance.customer_face_capture = customer_face_capture
            
        # Set the branch and created_by
        if self.request.user.branch:
            form.instance.branch = self.request.user.branch
        form.instance.created_by = self.request.user
        
        # Ensure interest_rate is set from the cleaned data
        if 'interest_rate' in form.cleaned_data and form.cleaned_data['interest_rate']:
            form.instance.interest_rate = form.cleaned_data['interest_rate']
        elif form.cleaned_data.get('scheme'):
            # Fallback: Set interest rate from scheme if not in cleaned data
            form.instance.interest_rate = form.cleaned_data['scheme'].interest_rate
        else:
            # Last fallback: Use default value from model
            form.instance.interest_rate = Decimal('12.00')
        
        messages.success(self.request, 'Loan created successfully!')
        return super().form_valid(form)


class LoanUpdateView(LoginRequiredMixin, RoleBranchAccessMixin, UpdateView):
    model = Loan
    form_class = LoanForm
    template_name = 'transactions/loan_form.html'
    slug_field = 'loan_number'
    slug_url_kwarg = 'loan_number'
    
    def get_success_url(self):
        return reverse('loan_detail', kwargs={'loan_number': self.object.loan_number})
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Process item photos for form editing
        if self.object and self.object.item_photos:
            context['item_photos_list'] = process_item_photos_for_display(self.object.item_photos)
        return context
    
    def form_valid(self, form):
        # Handle photo processing during loan update
        item_photos_data = self.request.POST.get('item_photos', '')
        customer_face_capture = self.request.POST.get('customer_face_capture', '')
        
        # Save photos directly to database
        if item_photos_data:
            form.instance.item_photos = item_photos_data
        if customer_face_capture:
            form.instance.customer_face_capture = customer_face_capture
        
        # Track who is editing the loan
        form.instance._edited_by = self.request.user
            
        messages.success(self.request, 'Loan updated successfully!')
        return super().form_valid(form)

    def get_object(self, queryset=None):
        obj = super().get_object(queryset=queryset)
        # enforce branch access
        self.check_object_branch_access(obj, branch_attr='branch')
        return obj

class LoanDeleteView(LoginRequiredMixin, ManagerPermissionMixin, RoleBranchAccessMixin, DeleteView):
    model = Loan
    template_name = 'transactions/loan_confirm_delete.html'
    slug_field = 'loan_number'
    slug_url_kwarg = 'loan_number'
    context_object_name = 'loan'
    success_url = reverse_lazy('loan_list')

    def delete(self, request, *args, **kwargs):
        self.object = self.get_object()
        loan_number = self.object.loan_number
        response = super().delete(request, *args, **kwargs)
        messages.success(request, f'Loan {loan_number} has been deleted successfully.')
        return response


class LoanEditLogsView(LoginRequiredMixin, RoleBranchAccessMixin, DetailView):
    """Display edit history for a loan"""
    model = Loan
    template_name = 'transactions/loan_edit_logs.html'
    slug_field = 'loan_number'
    slug_url_kwarg = 'loan_number'
    context_object_name = 'loan'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from accounts.models import LoanEditLog
        
        # Get all edit logs for this loan, ordered by date descending
        edit_logs = LoanEditLog.objects.filter(loan=self.object).order_by('-edited_at')
        context['edit_logs'] = edit_logs
        context['total_edits'] = edit_logs.count()
        
        return context

    def get_object(self, queryset=None):
        obj = super().get_object(queryset=queryset)
        # enforce branch access
        self.check_object_branch_access(obj, branch_attr='branch')
        return obj


class PaymentCreateView(LoginRequiredMixin, RoleBranchAccessMixin, CreateView):
    model = Payment
    template_name = 'transactions/payment_form.html'
    fields = ['amount', 'payment_date', 'payment_method', 'reference_number', 'notes']
    
    def get_initial(self):
        """Set default values for form fields"""
        initial = super().get_initial()
        # Set payment_date to today's date
        initial['payment_date'] = timezone.now().date()
        return initial
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        loan_number = self.kwargs.get('loan_number')
        loan = get_object_or_404(Loan, loan_number=loan_number)
        # enforce branch access
        self.check_object_branch_access(loan, branch_attr='branch')
        context['loan'] = loan
        
        # Calculate remaining balance for full payment
        try:
            remaining_balance = loan.total_payable_till_date - loan.amount_paid
            context['remaining_balance'] = max(remaining_balance, 0)
        except:
            context['remaining_balance'] = 0
            
        return context
    
    def form_valid(self, form):
        loan_number = self.kwargs.get('loan_number')
        loan = get_object_or_404(Loan, loan_number=loan_number)
        # enforce branch access
        self.check_object_branch_access(loan, branch_attr='branch')
        
        # Check if loan can accept payments
        if loan.status not in ['active', 'overdue']:
            messages.error(self.request, f'Cannot record payment for loan {loan_number}. Current status: {loan.get_status_display()}')
            return redirect('loan_detail', loan_number=loan_number)
        
        form.instance.loan = loan
        form.instance.received_by = self.request.user
        
        # Calculate remaining balance before this payment
        try:
            remaining_balance = loan.total_payable_till_date - loan.amount_paid
            payment_amount = form.instance.amount
            
            # Check if this payment will fully settle the loan
            will_close_loan = payment_amount >= remaining_balance
            
            # Save the payment first
            response = super().form_valid(form)
            
            # If payment fully settles the loan, close it
            if will_close_loan and remaining_balance > 0:
                loan.status = 'closed'
                loan.closure_date = timezone.now().date()
                loan.closed_by = self.request.user
                loan.save()
                
                messages.success(
                    self.request, 
                    f'Payment recorded successfully! Loan {loan_number} has been automatically closed as the full amount has been paid.'
                )
            else:
                messages.success(self.request, 'Payment recorded successfully!')
                
            return response
            
        except Exception as e:
            messages.error(self.request, f'Error processing payment: {str(e)}')
            return redirect('loan_detail', loan_number=loan_number)
    
    def get_success_url(self):
        return reverse('loan_detail', kwargs={'loan_number': self.kwargs.get('loan_number')})


class PaymentListView(LoginRequiredMixin, ListView):
    model = Payment
    template_name = 'transactions/interest_paid_list.html'
    context_object_name = 'payments'
    paginate_by = 20


class PaymentDetailView(LoginRequiredMixin, DetailView):
    model = Payment
    template_name = 'transactions/payment_detail.html'
    context_object_name = 'payment'


class LoanExtensionCreateView(LoginRequiredMixin, RoleBranchAccessMixin, CreateView):
    model = LoanExtension
    form_class = LoanExtensionForm
    template_name = 'transactions/loan_extension_form.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        loan_number = self.kwargs.get('loan_number')
        loan = get_object_or_404(Loan, loan_number=loan_number)
        # enforce branch access
        self.check_object_branch_access(loan, branch_attr='branch')
        context['loan'] = loan
        return context
    
    def form_valid(self, form):
        loan_number = self.kwargs.get('loan_number')
        loan = get_object_or_404(Loan, loan_number=loan_number)
        # enforce branch access
        self.check_object_branch_access(loan, branch_attr='branch')
        form.instance.loan = loan
        form.instance.created_by = self.request.user
        messages.success(self.request, 'Loan extension created successfully!')
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse('loan_detail', kwargs={'loan_number': self.kwargs.get('loan_number')})


class LoanForecloseView(LoginRequiredMixin, RoleBranchAccessMixin, View):
    def get(self, request, loan_number):
        """Handle GET request - show confirmation page"""
        loan = get_object_or_404(Loan, loan_number=loan_number)
        # enforce branch access
        self.check_object_branch_access(loan, branch_attr='branch')

        # Check if loan can be foreclosed
        if loan.status not in ['active', 'overdue']:
            messages.error(request, f'Loan {loan_number} cannot be foreclosed. Current status: {loan.get_status_display()}')
            return redirect('loan_detail', loan_number=loan_number)

        context = {
            'loan': loan,
            'confirm_action': 'foreclose'
        }
        return render(request, 'transactions/loan_foreclose_confirm.html', context)

    def post(self, request, loan_number):
        """Handle POST request - actually foreclose the loan"""
        loan = get_object_or_404(Loan, loan_number=loan_number)
        # enforce branch access
        self.check_object_branch_access(loan, branch_attr='branch')

        # Check if loan can be foreclosed
        if loan.status not in ['active', 'overdue']:
            messages.error(request, f'Loan {loan_number} cannot be foreclosed. Current status: {loan.get_status_display()}')
            return redirect('loan_detail', loan_number=loan_number)

        # Update loan status
        loan.status = 'foreclosed'
        loan.foreclosure_date = timezone.now().date()
        loan.foreclosed_by = request.user
        loan.save()

        messages.success(request, f'Loan {loan_number} has been successfully foreclosed.')
        return redirect('loan_detail', loan_number=loan_number)


class LoanDocumentView(LoginRequiredMixin, RoleBranchAccessMixin, View):
    def get(self, request, loan_number):
        loan = get_object_or_404(Loan, loan_number=loan_number)
        # enforce branch access
        self.check_object_branch_access(loan, branch_attr='branch')
        # Keep original agreement layout; prefer Chromium rendering for reliable Tamil glyph shaping.
        try:
            return self._generate_browser_pdf(request, loan)
        except Exception as e:
            print(f"Browser PDF generation failed, falling back to xhtml2pdf: {e}")
            return self._generate_xhtml2pdf(request, loan)

    def _find_browser_executable(self):
        """Find an installed Chromium-based browser executable."""
        candidates = [
            shutil.which('chrome'),
            shutil.which('msedge'),
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
            r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
        ]
        for candidate in candidates:
            if candidate and os.path.exists(candidate):
                return candidate
        return None

    def _generate_browser_pdf(self, request, loan):
        """Render loan agreement HTML via headless Chromium to get proper Tamil rendering."""
        current_language = getattr(request, 'LANGUAGE_CODE', 'en')
        # Process item photos for PDF using centralized function
        processed_photos = process_item_photos_for_display(loan.item_photos)
        item_photos = []
        for photo in processed_photos:
            if photo.startswith('data:image/'):
                base64_data = photo.split(',')[1] if ',' in photo else photo
                item_photos.append(base64_data)
            else:
                item_photos.append(photo)

        customer_photo = None
        if loan.customer_face_capture:
            if loan.customer_face_capture.startswith('data:image/'):
                customer_photo = loan.customer_face_capture.split(',')[1]
            else:
                customer_photo = loan.customer_face_capture

        loan_items = loan.loanitem_set.all()
        language_context = build_loan_pdf_language_context(loan, current_language)

        # Same filename logic used by template-based method
        from django.utils.text import slugify
        import re

        customer_name = ""
        if loan.customer:
            customer_name = f"{loan.customer.first_name}_{loan.customer.last_name}"
            customer_name = slugify(customer_name).replace('-', '_')

        item_names = []
        for loan_item in loan_items:
            if loan_item.item and loan_item.item.name:
                item_names.append(slugify(loan_item.item.name).replace('-', '_'))
        if not item_names:
            if hasattr(loan, 'item_name') and loan.item_name:
                item_names = [slugify(loan.item_name).replace('-', '_')]
            else:
                item_names = ['gold_item']
        items_part = '_'.join(item_names[:2])

        if customer_name and items_part:
            filename_base = f"{customer_name}_{items_part}_{loan.loan_number}_agreement"
        elif customer_name:
            filename_base = f"{customer_name}_{loan.loan_number}_agreement"
        else:
            filename_base = f"loan_{loan.loan_number}_agreement"
        filename_base = re.sub(r'[^a-zA-Z0-9_-]', '_', filename_base)[:200]
        detailed_filename = f"{filename_base}.pdf"

        context = {
            'loan': loan,
            'loan_items': loan_items,
            'item_photos': item_photos,
            'customer_photo': customer_photo,
            'tamil_font_file_uri': f"file:///{str((settings.BASE_DIR / 'static' / 'fonts' / 'NotoSansTamil-Regular.ttf')).replace(os.sep, '/')}",
            'pdf_renderer': 'browser',
            **language_context,
        }

        template = get_template('transactions/loan_document_pdf.html')
        html = template.render(context)

        browser = self._find_browser_executable()
        if not browser:
            raise RuntimeError("No Chrome/Edge executable found on system")

        tmp_dir = tempfile.mkdtemp(prefix='loan_pdf_')
        html_path = os.path.join(tmp_dir, 'loan_document.html')
        pdf_path = os.path.join(tmp_dir, 'loan_document.pdf')
        profile_dir = os.path.join(tmp_dir, 'profile')
        os.makedirs(profile_dir, exist_ok=True)

        try:
            with open(html_path, 'w', encoding='utf-8') as f:
                f.write(html)

            cmd = [
                browser,
                "--headless=new",
                "--disable-gpu",
                "--no-sandbox",
                f"--user-data-dir={profile_dir}",
                "--allow-file-access-from-files",
                "--disable-web-security",
                "--print-to-pdf-no-header",
                f"--print-to-pdf={pdf_path}",
                f"file:///{html_path.replace(os.sep, '/')}",
            ]

            result = subprocess.run(cmd, capture_output=True, text=True, timeout=90)
            if result.returncode != 0 or not os.path.exists(pdf_path):
                raise RuntimeError(f"Chromium PDF render failed: {result.stderr or result.stdout}")

            with open(pdf_path, 'rb') as f:
                pdf_bytes = f.read()
            if not pdf_bytes:
                raise RuntimeError("Generated PDF is empty")

            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{detailed_filename}"'
            response.write(pdf_bytes)
            return response
        finally:
            try:
                shutil.rmtree(tmp_dir, ignore_errors=True)
            except Exception:
                pass

    def _generate_xhtml2pdf(self, request, loan):
        """Fallback method using xhtml2pdf"""
        current_language = getattr(request, 'LANGUAGE_CODE', 'en')
        
        # Process item photos for PDF using centralized function
        processed_photos = process_item_photos_for_display(loan.item_photos)
        
        # Convert photos to base64 format for PDF embedding
        item_photos = []
        for photo in processed_photos:
            if photo.startswith('data:image/'):
                # Extract just the base64 data part (remove data:image/jpeg;base64, prefix)
                base64_data = photo.split(',')[1] if ',' in photo else photo
                item_photos.append(base64_data)
            else:
                # If it's already base64 without prefix, use directly
                item_photos.append(photo)
        
        # Debug: Print photo information
        print(f"Processing loan {loan.loan_number}: Found {len(processed_photos)} processed photos")
        print(f"Raw item_photos data: {loan.item_photos[:100] if loan.item_photos else 'None'}...")
        print(f"Final item_photos for template: {len(item_photos)} photos")
        
        # Process customer photo
        customer_photo = None
        if loan.customer_face_capture:
            if loan.customer_face_capture.startswith('data:image/'):
                customer_photo = loan.customer_face_capture.split(',')[1]
            else:
                customer_photo = loan.customer_face_capture
        
        # Ensure we have loan items
        loan_items = loan.loanitem_set.all()
        language_context = build_loan_pdf_language_context(loan, current_language)
        print(f"Found {loan_items.count()} loan items")
        
        # Generate detailed filename with customer name and item details
        def generate_loan_document_filename(loan):
            from django.utils.text import slugify
            import re
            
            # Get customer name (clean it for filename)
            customer_name = ""
            if loan.customer:
                customer_name = f"{loan.customer.first_name}_{loan.customer.last_name}"
                customer_name = slugify(customer_name).replace('-', '_')
            
            # Get item names from loan items
            item_names = []
            loan_items = loan.loanitem_set.all()
            for loan_item in loan_items:
                if loan_item.item and loan_item.item.name:
                    item_name = slugify(loan_item.item.name).replace('-', '_')
                    item_names.append(item_name)
            
            # If no items found, use item_name from loan model or default
            if not item_names:
                if hasattr(loan, 'item_name') and loan.item_name:
                    item_name = slugify(loan.item_name).replace('-', '_')
                    item_names = [item_name]
                else:
                    item_names = ['gold_item']
            
            # Combine item names (limit to first 2 items to avoid very long filenames)
            items_part = '_'.join(item_names[:2])
            
            # Create filename: CustomerName_ItemNames_LoanNumber_agreement.pdf
            if customer_name and items_part:
                filename_base = f"{customer_name}_{items_part}_{loan.loan_number}_agreement"
            elif customer_name:
                filename_base = f"{customer_name}_{loan.loan_number}_agreement"
            else:
                filename_base = f"loan_{loan.loan_number}_agreement"
            
            # Clean filename for filesystem compatibility
            filename_base = re.sub(r'[^a-zA-Z0-9_-]', '_', filename_base)
            
            # Limit filename length to avoid filesystem issues
            if len(filename_base) > 200:
                filename_base = filename_base[:200]
            
            return f"{filename_base}.pdf"
        
        context = {
            'loan': loan,
            'loan_items': loan_items,
            'item_photos': item_photos,
            'customer_photo': customer_photo,
            'pdf_renderer': 'xhtml2pdf',
            **language_context,
        }
        
        # Render PDF
        template = get_template('transactions/loan_document_pdf.html')
        html = template.render(context)
        
        response = HttpResponse(content_type='application/pdf')
        
        # Generate detailed filename
        detailed_filename = generate_loan_document_filename(loan)
        response['Content-Disposition'] = f'attachment; filename="{detailed_filename}"'
        
        def link_callback(uri, rel):
            """
            Resolve static/media URIs to absolute filesystem paths for xhtml2pdf.
            Required so Tamil font files under /static/fonts can be loaded.
            """
            parsed = urlparse(uri)
            path = parsed.path or uri
            if path.startswith(settings.STATIC_URL):
                return os.path.join(settings.BASE_DIR, 'static', path.replace(settings.STATIC_URL, '', 1))
            if path.startswith(settings.MEDIA_URL):
                return os.path.join(settings.MEDIA_ROOT, path.replace(settings.MEDIA_URL, '', 1))
            if path.startswith('/static/'):
                return os.path.join(settings.BASE_DIR, 'static', path.replace('/static/', '', 1))
            if path.startswith('/media/'):
                return os.path.join(settings.MEDIA_ROOT, path.replace('/media/', '', 1))
            return uri

        # Register Tamil font directly with reportlab to avoid xhtml2pdf @font-face temp-file issues on Windows.
        try:
            tamil_font_path = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'NotoSansTamil-Regular.ttf')
            if os.path.exists(tamil_font_path):
                pdfmetrics.registerFont(TTFont('NotoSansTamil', tamil_font_path))
        except Exception:
            pass

        pisa_status = pisa.CreatePDF(html, dest=response, link_callback=link_callback)
        if pisa_status.err:
            return HttpResponse('Error generating PDF', status=500)
        
        return response


class LoanPaymentHistoryDownloadView(LoginRequiredMixin, RoleBranchAccessMixin, View):
    def get(self, request, loan_number):
        loan = get_object_or_404(Loan, loan_number=loan_number)
        # enforce branch access
        self.check_object_branch_access(loan, branch_attr='branch')
        format_type = request.GET.get('format', 'csv')
        
        if format_type == 'csv':
            return self.export_csv(loan)
        elif format_type == 'excel':
            return self.export_excel(loan)
        elif format_type == 'pdf':
            return self.export_pdf(loan)
        else:
            return self.export_csv(loan)
    
    def export_csv(self, loan):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="loan_{loan.loan_number}_payment_history.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Date', 'Amount', 'Method', 'Reference', 'Notes'])
        
        for payment in loan.payments.all().order_by('-payment_date'):
            writer.writerow([
                payment.payment_date.strftime('%Y-%m-%d'),
                payment.amount,
                payment.get_payment_method_display(),
                payment.reference_number or '',
                payment.notes or ''
            ])
        
        return response


class LoanScheduleView(LoginRequiredMixin, RoleBranchAccessMixin, View):
    """Generate a month-by-month interest schedule from current month to due month.
    CSV download is available via ?download=csv
    """
    def get(self, request, loan_number):
        loan = get_object_or_404(Loan, loan_number=loan_number)
        # enforce branch access
        self.check_object_branch_access(loan, branch_attr='branch')

        # Basic inputs (use Decimal for precise currency arithmetic)
        # Principal should be Distribution Amount + Processing Fee when available
        dist_amt = getattr(loan, 'distribution_amount', None)
        proc_fee = getattr(loan, 'processing_fee', None)
        try:
            dist_amt_d = Decimal(str(dist_amt)) if dist_amt is not None else Decimal('0')
        except Exception:
            dist_amt_d = Decimal('0')
        try:
            proc_fee_d = Decimal(str(proc_fee)) if proc_fee is not None else Decimal('0')
        except Exception:
            proc_fee_d = Decimal('0')

        if dist_amt_d > Decimal('0'):
            principal = (dist_amt_d + proc_fee_d).quantize(Decimal('0.01'))
        else:
            principal = Decimal(str(getattr(loan, 'principal_amount', 0) or 0))

        rate_annual = Decimal(str(getattr(loan, 'interest_rate', 0) or 0))
        today = timezone.now().date()
        due = getattr(loan, 'due_date', None)
        if not due or principal <= 0 or rate_annual <= 0:
            return HttpResponse('Loan missing principal, due date, or interest rate', status=400)

        # start from beginning of current month
        start = today.replace(day=1)
        # iterate months until due month inclusive
        months = []
        cur = start
        while cur <= due:
            months.append(cur)
            # advance to next month
            year = cur.year + (cur.month // 12)
            month = (cur.month % 12) + 1
            cur = cur.replace(year=year, month=month, day=1)

        monthly_rate = (rate_annual / Decimal('100')) / Decimal('12')

        rows = []
        remaining_principal = principal

        num_months = max(1, len(months))
        # Distribute principal evenly using Decimal; adjust last month to absorb rounding
        principal_per_month = (principal / Decimal(num_months)).quantize(Decimal('0.01'))
        last_month_principal = (principal - principal_per_month * (num_months - 1)).quantize(Decimal('0.01'))

        # Interest is constant each month based on original principal
        monthly_interest_const = (principal * monthly_rate).quantize(Decimal('0.01'))

        # Log diagnostics for inspection
        try:
            logger.info(
                "LoanSchedule diagnostics: loan=%s principal=%s distribution_amount=%s processing_fee=%s num_months=%s principal_per_month=%s last_month_principal=%s monthly_interest_const=%s",
                loan.loan_number,
                str(principal),
                str(dist_amt_d),
                str(proc_fee_d),
                str(num_months),
                str(principal_per_month),
                str(last_month_principal),
                str(monthly_interest_const),
            )
        except Exception:
            pass

        # Build rows: principal shown as full principal (distribution+processing_fee) each month,
        # interest constant, total = principal + cumulative interest up to that month.
        for idx, m in enumerate(months):
            month_index = Decimal(idx + 1)
            cumulative_interest = (monthly_interest_const * month_index).quantize(Decimal('0.01'))
            total_amount = (principal + cumulative_interest).quantize(Decimal('0.01'))

            # Prepare principal display with breakdown e.g., 35354(35000+354)
            dist_display = dist_amt_d.quantize(Decimal('0.01'))
            proc_display = proc_fee_d.quantize(Decimal('0.01'))
            # Format amounts as integer if no cents, else two decimals
            def fmt(a: Decimal):
                a_q = a.quantize(Decimal('0.01'))
                if a_q == a_q.to_integral():
                    return str(int(a_q))
                return format(a_q, '0.2f')

            principal_display = f"{fmt(principal)}({fmt(dist_display)}+{fmt(proc_display)})"

            rows.append({
                'month': m.strftime('%b-%Y'),
                'principal': principal,
                'principal_display': principal_display,
                'interest': monthly_interest_const,
                'total_amount': total_amount,
            })

        # final row: totals
        total_interest = (monthly_interest_const * Decimal(num_months)).quantize(Decimal('0.01'))
        final_row = {
            'month': 'TOTAL',
            'interest': total_interest,
            'principal': principal.quantize(Decimal('0.01')),
            'total_amount': (principal + total_interest).quantize(Decimal('0.01')),
            'payoff_amount': (principal + Decimal('0.00') + monthly_interest_const).quantize(Decimal('0.01')),
        }

        download = request.GET.get('download')
        if download == 'csv':
            # return CSV with new columns
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="loan_{loan.loan_number}_schedule.csv"'
            writer = csv.writer(response)
            writer.writerow(['Month', 'Principal', 'Monthly Interest', 'Total Amount'])
            for idx, r in enumerate(rows, start=1):
                # cumulative interest = monthly_interest * idx
                cum_interest = (r['interest'] * Decimal(idx)).quantize(Decimal('0.01'))
                writer.writerow([r['month'], r.get('principal_display') or format(r['principal'], '0.2f'), format(r['interest'], '0.2f'), format(r['total_amount'], '0.2f')])
            writer.writerow([final_row['month'], format(final_row['principal'], '0.2f'), format(final_row['interest'], '0.2f'), format(final_row['total_amount'], '0.2f')])
            return response

        if download == 'pdf':
            # Generate a simple PDF schedule using ReportLab
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch

            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="loan_{loan.loan_number}_schedule.pdf"'

            doc = SimpleDocTemplate(
                response,
                pagesize=landscape(A4),
                rightMargin=0.3*inch,
                leftMargin=0.3*inch,
                topMargin=0.5*inch,
                bottomMargin=0.5*inch,
            )

            styles = getSampleStyleSheet()
            title_style = ParagraphStyle('Title', parent=styles['Heading2'], alignment=1)
            normal = styles['Normal']

            elements = []
            elements.append(Paragraph(f"Loan Schedule - {loan.loan_number}", title_style))
            elements.append(Spacer(1, 12))

            table_data = [[ 'Month', 'Principal', 'Monthly Interest', 'Total Amount' ]]
            for idx, r in enumerate(rows, start=1):
                table_data.append([r['month'], r.get('principal_display') or format(r['principal'], '0.2f'), f"{r['interest']:.2f}", f"{r['total_amount']:.2f}"])
            table_data.append([final_row['month'], format(final_row['principal'], '0.2f'), f"{final_row['interest']:.2f}", f"{final_row['total_amount']:.2f}"])

            col_widths = [1.4*inch, 2.0*inch, 1.5*inch, 1.8*inch]
            table = Table(table_data, colWidths=col_widths)

            # Build table style and add alternating row backgrounds for readability
            style_list = [
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#366092')),
                ('TEXTCOLOR',(0,0),(-1,0),colors.white),
                ('ALIGN',(1,1),(-1,-1),'RIGHT'),
                ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
                ('GRID',(0,0),(-1,-1),0.5,colors.grey),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ]

            # Apply zebra striping to data rows (row index starts at 0 for header)
            for row_idx in range(1, len(table_data)):
                if row_idx % 2 == 0:
                    # even data rows -> light blue background
                    style_list.append(('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#f6f9ff')))

            table.setStyle(TableStyle(style_list))

            elements.append(table)
            doc.build(elements)
            return response

        context = {
            'loan': loan,
            'rows': rows,
            'final_row': final_row,
            'diagnostics': {
                'principal': str(principal),
                'distribution_amount': str(dist_amt_d),
                'processing_fee': str(proc_fee_d),
                'num_months': num_months,
                'principal_per_month': str(principal_per_month),
                'last_month_principal': str(last_month_principal),
                'monthly_interest_const': str(monthly_interest_const),
            },
        }
        return render(request, 'transactions/loan_schedule.html', context)
    
    def export_excel(self, loan):
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = f"Loan {loan.loan_number} Payments"
        
        # Headers
        headers = ['Date', 'Amount', 'Method', 'Reference', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        for row_idx, payment in enumerate(loan.payments.all().order_by('-payment_date'), 2):
            worksheet.cell(row=row_idx, column=1, value=payment.payment_date.strftime('%Y-%m-%d'))
            worksheet.cell(row=row_idx, column=2, value=float(payment.amount))
            worksheet.cell(row=row_idx, column=3, value=payment.get_payment_method_display())
            worksheet.cell(row=row_idx, column=4, value=payment.reference_number or '')
            worksheet.cell(row=row_idx, column=5, value=payment.notes or '')
        
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="loan_{loan.loan_number}_payment_history.xlsx"'
        
        return response
    
    def export_pdf(self, loan):
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from datetime import datetime
        import os
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="payment_history_{loan.loan_number}_{datetime.now().strftime("%Y%m%d")}.pdf"'
        
        # Create PDF document with better margins
        doc = SimpleDocTemplate(
            response, 
            pagesize=A4,
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.75*inch,
            bottomMargin=0.75*inch
        )
        elements = []
        styles = getSampleStyleSheet()
        
        # Register Unicode font for better currency symbol support
        try:
            possible_fonts = [
                '/System/Library/Fonts/Arial.ttf',  # macOS
                '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',  # Linux
                'C:/Windows/Fonts/arial.ttf',  # Windows
            ]
            
            for font_path in possible_fonts:
                if os.path.exists(font_path):
                    pdfmetrics.registerFont(TTFont('CustomFont', font_path))
                    custom_font = 'CustomFont'
                    break
            else:
                custom_font = 'Helvetica'
        except:
            custom_font = 'Helvetica'
        
        # Create custom styles
        company_style = ParagraphStyle(
            'CompanyStyle',
            parent=styles['Normal'],
            fontSize=16,
            fontName='Helvetica-Bold',
            alignment=1,  # Center
            spaceAfter=5,
            textColor=colors.HexColor('#2C3E50')
        )
        
        address_style = ParagraphStyle(
            'AddressStyle',
            parent=styles['Normal'],
            fontSize=10,
            fontName=custom_font,
            alignment=1,  # Center
            spaceAfter=10,
            textColor=colors.HexColor('#7F8C8D')
        )
        
        title_style = ParagraphStyle(
            'TitleStyle',
            parent=styles['Heading1'],
            fontSize=18,
            fontName='Helvetica-Bold',
            alignment=1,  # Center
            spaceAfter=20,
            textColor=colors.HexColor('#34495E'),
            borderWidth=2,
            borderColor=colors.HexColor('#3498DB'),
            borderPadding=10,
            backColor=colors.HexColor('#ECF0F1')
        )
        
        section_style = ParagraphStyle(
            'SectionStyle',
            parent=styles['Normal'],
            fontSize=12,
            fontName='Helvetica-Bold',
            spaceAfter=10,
            textColor=colors.HexColor('#2C3E50')
        )
        
        info_style = ParagraphStyle(
            'InfoStyle',
            parent=styles['Normal'],
            fontSize=10,
            fontName=custom_font,
            spaceAfter=5,
            textColor=colors.HexColor('#34495E')
        )
        
        # Company Header
        company_name = "PAWNSHOP MANAGEMENT SYSTEM"
        
        # Get branch details if available
        if loan.branch:
            branch_info = loan.branch
            company_name = f"{branch_info.name.upper()}"
            
            # Add company/branch name
            elements.append(Paragraph(company_name, company_style))
            
            # Add branch address if available
            address_parts = []
            if hasattr(branch_info, 'address') and branch_info.address:
                address_parts.append(branch_info.address)
            if hasattr(branch_info, 'city') and branch_info.city:
                address_parts.append(branch_info.city)
            if hasattr(branch_info, 'state') and branch_info.state:
                address_parts.append(branch_info.state)
            if hasattr(branch_info, 'pincode') and branch_info.pincode:
                address_parts.append(f"PIN: {branch_info.pincode}")
                
            if address_parts:
                elements.append(Paragraph(", ".join(address_parts), address_style))
            
            # Add contact details
            contact_parts = []
            branch_header_phones = get_branch_bill_header_phones(branch_info)
            if branch_header_phones:
                contact_parts.append(f"Phone: {branch_header_phones}")
            if hasattr(branch_info, 'email') and branch_info.email:
                contact_parts.append(f"Email: {branch_info.email}")
                
            if contact_parts:
                elements.append(Paragraph(" | ".join(contact_parts), address_style))
        else:
            elements.append(Paragraph(company_name, company_style))
            elements.append(Paragraph("Professional Pawnshop Services", address_style))
        
        # Add horizontal line
        elements.append(Spacer(1, 10))
        
        # Document Title
        elements.append(Paragraph("PAYMENT HISTORY REPORT", title_style))
        elements.append(Spacer(1, 20))
        
        # Loan Information Section
        elements.append(Paragraph("LOAN INFORMATION", section_style))
        
        # Create loan info table
        loan_data = [
            ['Loan Number:', loan.loan_number],
            ['Customer Name:', f"{loan.customer.first_name} {loan.customer.last_name}"],
            ['Customer Phone:', getattr(loan.customer, 'phone', 'N/A')],
            ['Principal Amount:', f"Rs {loan.principal_amount:,.2f}"],
            ['Issue Date:', loan.issue_date.strftime('%d %B %Y')],
            ['Due Date:', loan.due_date.strftime('%d %B %Y')],
            ['Loan Status:', loan.get_status_display()],
        ]
        
        loan_info_table = Table(loan_data, colWidths=[2*inch, 4*inch])
        loan_info_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), custom_font),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BDC3C7')),
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#ECF0F1')),
        ]))
        
        elements.append(loan_info_table)
        elements.append(Spacer(1, 20))
        
        # Payment Summary
        total_payments = loan.payments.count()
        total_amount_paid = sum(payment.amount for payment in loan.payments.all())
        
        elements.append(Paragraph("PAYMENT SUMMARY", section_style))
        
        # Calculate remaining balance - ensure it's never negative (0 for fully paid loans)
        remaining_balance = max(0, loan.total_payable_till_date - total_amount_paid)
        
        summary_data = [
            ['Total Payments Made:', str(total_payments)],
            ['Total Amount Paid:', f"Rs {total_amount_paid:,.2f}"],
            ['Remaining Balance:', f"Rs {remaining_balance:,.2f}"],
        ]
        
        summary_table = Table(summary_data, colWidths=[2*inch, 4*inch])
        summary_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (1, 0), (1, -1), custom_font),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BDC3C7')),
            ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E8F6F3')),
            ('BACKGROUND', (1, -1), (1, -1), colors.HexColor('#FADBD8')),  # Highlight remaining balance
        ]))
        
        elements.append(summary_table)
        elements.append(Spacer(1, 20))
        
        # Payment Details Section
        elements.append(Paragraph("PAYMENT DETAILS", section_style))
        
        if loan.payments.exists():
            # Payment history table headers
            payment_headers = ['S.No.', 'Date', 'Amount (Rs)', 'Method', 'Reference No.', 'Received By', 'Notes']
            payment_data = [payment_headers]
            
            # Add payment rows
            for idx, payment in enumerate(loan.payments.all().order_by('-payment_date'), 1):
                row = [
                    str(idx),
                    payment.payment_date.strftime('%d-%m-%Y'),
                    f"{payment.amount:,.2f}",
                    payment.get_payment_method_display(),
                    payment.reference_number or '-',
                    f"{payment.received_by.first_name} {payment.received_by.last_name}" if payment.received_by else 'N/A',
                    payment.notes[:30] + '...' if payment.notes and len(payment.notes) > 30 else (payment.notes or '-')
                ]
                payment_data.append(row)
            
            payment_table = Table(payment_data, colWidths=[0.5*inch, 1*inch, 1*inch, 1*inch, 1*inch, 1.2*inch, 1.3*inch])
            payment_table.setStyle(TableStyle([
                # Header style
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                
                # Data rows style
                ('FONTNAME', (0, 1), (-1, -1), custom_font),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # S.No center
                ('ALIGN', (1, 1), (1, -1), 'CENTER'),  # Date center
                ('ALIGN', (2, 1), (2, -1), 'RIGHT'),   # Amount right
                ('ALIGN', (3, 1), (-1, -1), 'LEFT'),   # Rest left
                
                # General styling
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                
                # Grid and alternating colors
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BDC3C7')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
                
                # Amount column highlighting
                ('BACKGROUND', (2, 1), (2, -1), colors.HexColor('#E8F8F5')),
            ]))
            
            elements.append(payment_table)
        else:
            elements.append(Paragraph("No payments recorded for this loan.", info_style))
        
        elements.append(Spacer(1, 30))
        
        # Footer section
        footer_style = ParagraphStyle(
            'FooterStyle',
            parent=styles['Normal'],
            fontSize=8,
            fontName=custom_font,
            alignment=1,  # Center
            textColor=colors.HexColor('#7F8C8D')
        )
        
        # Add generation info
        elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%d %B %Y at %I:%M %p')}", footer_style))
        elements.append(Paragraph("This is a computer-generated document and does not require a signature.", footer_style))
        
        # Add disclaimer
        elements.append(Spacer(1, 10))
        disclaimer_style = ParagraphStyle(
            'DisclaimerStyle',
            parent=styles['Normal'],
            fontSize=7,
            fontName=custom_font,
            alignment=4,  # Justify
            textColor=colors.HexColor('#95A5A6'),
            leftIndent=20,
            rightIndent=20
        )
        
        disclaimer_text = ("This payment history is provided for informational purposes only. "
                          "All payment details are subject to verification. For any discrepancies, "
                          "please contact the branch office immediately.")
        elements.append(Paragraph(disclaimer_text, disclaimer_style))
        
        # Build PDF
        doc.build(elements)
        return response


class PaymentReceiptView(LoginRequiredMixin, View):
    def get(self, request, payment_id):
        payment = get_object_or_404(Payment, id=payment_id)
        loan = payment.loan
        bill_details = get_branch_bill_details(getattr(loan, 'branch', None))
        
        # Prepare amount in words (English) and Tamil
        amount_in_words = amount_to_english_words(payment.amount)
        amount_in_words_tamil = number_to_tamil_words(payment.amount)

        # Keep receipt generation fast by avoiding runtime network translation.
        payment.notes_tamil = payment.notes or ''

        context = {
            'payment': payment,
            'loan': loan,
            'branch_phone_display': get_branch_bill_header_phones(getattr(loan, 'branch', None)),
            'branch_address_display': bill_details.get('address', ''),
            'bill_shop_name': bill_details.get('shop_name', ''),
            'bill_address_display': bill_details.get('address', ''),
            'bill_phone_display': bill_details.get('phone', ''),
            'bill_email_display': bill_details.get('email', ''),
            'bill_logo_url': bill_details.get('logo_url', ''),
            'total_items_count': get_loan_total_items_count(loan),
            'amount_in_words': amount_in_words,
            'amount_in_words_tamil': amount_in_words_tamil,
            'tamil_font_file_uri': f"file:///{str((settings.BASE_DIR / 'static' / 'fonts' / 'NotoSansTamil-Regular.ttf')).replace(os.sep, '/')}",
        }
        
        # Render PDF
        template = get_template('transactions/payment_receipt_pdf.html')
        html = template.render(context)
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="payment_receipt_{payment.id}.pdf"'
        
        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            return HttpResponse('Error generating PDF', status=500)
        
        return response


class SaleListView(LoginRequiredMixin, ListView):
    model = Sale
    template_name = 'transactions/sale_list.html'
    context_object_name = 'sales'
    paginate_by = 20


class SaleCreateView(LoginRequiredMixin, CreateView):
    model = Sale
    form_class = SaleForm
    template_name = 'transactions/sale_form.html'
    success_url = reverse_lazy('sale_list')
    
    def form_valid(self, form):
        form.instance.created_by = self.request.user
        if self.request.user.branch:
            form.instance.branch = self.request.user.branch
        messages.success(self.request, 'Sale created successfully!')
        return super().form_valid(form)


class SaleDetailView(LoginRequiredMixin, DetailView):
    model = Sale
    template_name = 'transactions/sale_detail.html'
    context_object_name = 'sale'


class SaleUpdateView(LoginRequiredMixin, UpdateView):
    model = Sale
    form_class = SaleForm
    template_name = 'transactions/sale_form.html'
    
    def get_success_url(self):
        return reverse('sale_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        messages.success(self.request, 'Sale updated successfully!')
        return super().form_valid(form)


class SaleCancelView(LoginRequiredMixin, View):
    def post(self, request, pk):
        sale = get_object_or_404(Sale, pk=pk)
        sale.status = 'cancelled'
        sale.save()
        messages.success(request, f'Sale #{sale.id} has been cancelled.')
        return redirect('sale_detail', pk=pk)


class SaleCompleteView(LoginRequiredMixin, View):
    def post(self, request, pk):
        sale = get_object_or_404(Sale, pk=pk)
        sale.status = 'completed'
        sale.save()
        messages.success(request, f'Sale #{sale.id} has been completed.')
        return redirect('sale_detail', pk=pk)


class SaleReceiptView(LoginRequiredMixin, View):
    def get(self, request, pk):
        sale = get_object_or_404(Sale, pk=pk)
        bill_details = get_branch_bill_details(getattr(sale, 'branch', None))
        
        # Compute English and Tamil amount-in-words for the sale total
        try:
            total_amount = sale.total_amount
        except Exception:
            total_amount = getattr(sale, 'total_amount', 0)

        sale.total_amount_in_words = amount_to_english_words(total_amount)
        sale.total_amount_in_words_tamil = number_to_tamil_words(total_amount)

        context = {
            'sale': sale,
            'branch_phone_display': get_branch_bill_header_phones(getattr(sale, 'branch', None)),
            'branch_address_display': bill_details.get('address', ''),
            'bill_shop_name': bill_details.get('shop_name', ''),
            'bill_address_display': bill_details.get('address', ''),
            'bill_phone_display': bill_details.get('phone', ''),
            'bill_email_display': bill_details.get('email', ''),
            'bill_logo_url': bill_details.get('logo_url', ''),
            'total_items_count': 1,
            'now': timezone.now(),
            'tamil_font_file_uri': f"file:///{str((settings.BASE_DIR / 'static' / 'fonts' / 'NotoSansTamil-Regular.ttf')).replace(os.sep, '/')}",
        }
        
        # Render PDF
        template = get_template('transactions/sale_receipt_pdf.html')
        html = template.render(context)
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="sale_receipt_{sale.id}.pdf"'
        
        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            return HttpResponse('Error generating PDF', status=500)
        
        return response


def number_to_words(request, number):
    """Utility view to convert numbers to words"""
    try:
        words = num2words(float(number))
        return JsonResponse({'words': words})
    except (ValueError, TypeError):
        return JsonResponse({'words': 'Invalid number'}, status=400)

